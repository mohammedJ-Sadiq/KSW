# -*- coding: utf-8 -*-
"""Tests for the unified bank file export wizard (ksw.bank.file.export.wizard).

Covers:
  - Wizard field defaults & constraints
  - action_open_export_wizard from batch
  - Export mode: all_excel, all_txt, specific_excel, specific_txt
  - WPS employees scoped to WPS bank only (no Kawthar employees)
  - Kawthar employees scoped to Kawthar bank only
  - Kawthar TXT generation: 194-char fixed-width lines, field positions
  - _bank_label uses acc_number for unique filenames
  - _bundle_and_download: single file vs. zip bundle
  - Error handling: no slips, no bank, no file type, missing value date
"""
import base64
import io
import zipfile
from datetime import date

from odoo.exceptions import UserError
from odoo.tests.common import TransactionCase

try:
    import openpyxl
except ImportError:
    openpyxl = None


class TestBankFileExportWizard(TransactionCase):
    """Tests for ksw.bank.file.export.wizard."""

    KAWTHAR_LINE_LEN = 194

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        cls.company = cls.env.company
        cls.company_partner = cls.company.partner_id

        # ── Banks ──
        cls.bank_rajhi = cls.env['res.bank'].create({
            'name': 'Al Rajhi Bank',
            'bic': 'RJHISARI',
        })
        cls.bank_anb = cls.env['res.bank'].create({
            'name': 'Arab National Bank',
            'bic': 'ARNBSARI',
        })

        # ── Company bank accounts ──
        # acc_number is unique per partner, and the dev/production databases
        # hold real company accounts with these names — hence the TEST prefix.
        # WPS bank account
        cls.wps_bank = cls.env['res.partner.bank'].create({
            'acc_number': 'TEST WPS Main Account',
            'partner_id': cls.company_partner.id,
            'bank_id': cls.bank_anb.id,
            'x_wps_cic_number': '1389678',
            'x_wps_debit_account': 'SA7180000578608010033217',
            'x_wps_mol_id': '004-110457',
            'x_file_type': 'wps',
        })
        # Kawthar bank accounts (two, to test unique filenames)
        cls.kawthar_bank_1 = cls.env['res.partner.bank'].create({
            'acc_number': 'TEST Kawther Raj Cards',
            'partner_id': cls.company_partner.id,
            'bank_id': cls.bank_rajhi.id,
            'x_wps_cic_number': '5555555',
            'x_file_type': 'kawthar',
        })
        cls.kawthar_bank_2 = cls.env['res.partner.bank'].create({
            'acc_number': 'TEST Hayat Raj Cards',
            'partner_id': cls.company_partner.id,
            'bank_id': cls.bank_rajhi.id,
            'x_wps_cic_number': '6666666',
            'x_file_type': 'kawthar',
        })

        # ── Work schedule ──
        cls.calendar = cls.env['resource.calendar'].create({
            'name': 'Export Test Calendar',
            'tz': 'Asia/Riyadh',
        })

        # ── Salary rule refs ──
        cls.cat_basic = cls.env.ref('om_hr_payroll.BASIC')
        cls.cat_hra = cls.env.ref('om_hr_payroll.HRA')
        cls.cat_gross = cls.env.ref('om_hr_payroll.GROSS')
        cls.cat_ded = cls.env.ref('om_hr_payroll.DED')
        cls.cat_net = cls.env.ref('om_hr_payroll.NET')

        cls.rule_basic = cls._get_or_create_rule(cls, 'BASIC', cls.cat_basic)
        cls.rule_hra = cls._get_or_create_rule(cls, 'HRA', cls.cat_hra)
        cls.rule_gross = cls._get_or_create_rule(cls, 'GROSS', cls.cat_gross)
        cls.rule_ded = cls._get_or_create_rule(cls, 'ATTDED', cls.cat_ded)
        cls.rule_net = cls._get_or_create_rule(cls, 'NET', cls.cat_net)

        # ── Employees ──
        # WPS employee
        cls.emp_wps = cls._create_employee(
            cls, 'AHMED WPS EMPLOYEE', '101', '1001001001',
            cls.bank_anb, 'SA11111111111111111W0001',
        )
        cls.emp_wps.sudo().write({
            'x_salary_bank_account_id': cls.wps_bank.id,
        })

        # Kawthar employee (bank 1)
        cls.emp_kaw_1 = cls._create_employee(
            cls, 'ANITA KAWTHAR ONE', '201', '2002002001',
            cls.bank_rajhi, '5689110000130257800',
        )
        cls.emp_kaw_1.sudo().write({
            'x_salary_bank_account_id': cls.kawthar_bank_1.id,
        })

        # Kawthar employee (bank 2)
        cls.emp_kaw_2 = cls._create_employee(
            cls, 'OLIVER KAWTHAR TWO', '202', '2002002002',
            cls.bank_rajhi, '5689110000105257800',
        )
        cls.emp_kaw_2.sudo().write({
            'x_salary_bank_account_id': cls.kawthar_bank_2.id,
        })

        # ── Payslip batch ──
        cls.batch = cls.env['hr.payslip.run'].create({
            'name': 'March 2026',
            'date_start': date(2026, 3, 1),
            'date_end': date(2026, 3, 31),
        })

        # ── Payslips ──
        cls.slip_wps = cls._create_payslip(
            cls, cls.emp_wps, cls.batch,
            basic=8000, hra=2000, gross=12000, deductions=1000, net=11000,
        )
        cls.slip_kaw_1 = cls._create_payslip(
            cls, cls.emp_kaw_1, cls.batch,
            basic=5000, hra=1000, gross=7000, deductions=500, net=6500,
        )
        cls.slip_kaw_2 = cls._create_payslip(
            cls, cls.emp_kaw_2, cls.batch,
            basic=6000, hra=0, gross=6000, deductions=0, net=6000,
        )

    # ── Factory helpers ──

    def _get_or_create_rule(self, code, category):
        rule = self.env['hr.salary.rule'].search(
            [('code', '=', code)], limit=1,
        )
        if not rule:
            rule = self.env['hr.salary.rule'].create({
                'name': code,
                'code': code,
                'category_id': category.id,
                'sequence': 100,
                'condition_select': 'none',
                'amount_select': 'fix',
                'amount_fix': 0,
            })
        return rule

    def _create_employee(self, name, barcode, id_number, bank, iban):
        emp = self.env['hr.employee'].create({
            'name': name,
            'barcode': barcode,
            'identification_id': id_number,
            'resource_calendar_id': self.calendar.id,
        })
        if iban:
            bank_acc = self.env['res.partner.bank'].create({
                'acc_number': iban,
                'partner_id': emp.work_contact_id.id,
                'bank_id': bank.id,
            })
            emp.sudo().write({'bank_account_ids': [(4, bank_acc.id)]})
        return emp

    def _create_payslip(self, employee, batch, basic, hra, gross,
                        deductions, net):
        slip = self.env['hr.payslip'].create({
            'name': 'Slip %s' % employee.name,
            'employee_id': employee.id,
            'date_from': batch.date_start,
            'date_to': batch.date_end,
            'payslip_run_id': batch.id,
            'version_id': employee.current_version_id.id,
        })
        Line = self.env['hr.payslip.line']
        version_id = employee.current_version_id.id

        def _mk(rule, amount):
            Line.create({
                'slip_id': slip.id,
                'salary_rule_id': rule.id,
                'code': rule.code,
                'name': rule.name,
                'category_id': rule.category_id.id,
                'employee_id': employee.id,
                'version_id': version_id,
                'amount': amount,
                'quantity': 1,
                'rate': 100,
                'sequence': rule.sequence,
            })

        _mk(self.rule_basic, basic)
        _mk(self.rule_hra, hra)
        _mk(self.rule_gross, gross)
        if deductions:
            _mk(self.rule_ded, -abs(deductions))
        _mk(self.rule_net, net)
        return slip

    def _make_wizard(self, batch=None, mode='all_excel', bank=None,
                     value_date=None, operation_code='2'):
        """Create a unified export wizard instance."""
        vals = {
            'payslip_run_id': (batch or self.batch).id,
            'export_mode': mode,
            'value_date': value_date or date(2026, 3, 30),
            'operation_code': operation_code,
        }
        if bank:
            vals['bank_account_id'] = bank.id
        return self.env['ksw.bank.file.export.wizard'].create(vals)

    # ================================================================
    # Tests — Wizard basics
    # ================================================================

    def test_wizard_defaults(self):
        """Wizard has correct defaults."""
        wiz = self.env['ksw.bank.file.export.wizard'].with_context(
            default_payslip_run_id=self.batch.id,
        ).create({})
        self.assertEqual(wiz.export_mode, 'all_excel')
        self.assertEqual(wiz.operation_code, '2')
        self.assertTrue(wiz.value_date)

    def test_company_partner_id_computed(self):
        """company_partner_id uses env.company, not payslip_run_id.company_id."""
        wiz = self._make_wizard()
        self.assertEqual(wiz.company_partner_id, self.env.company.partner_id)

    def test_action_open_export_wizard(self):
        """Batch button returns act_window for the unified wizard."""
        action = self.batch.action_open_export_wizard()
        self.assertEqual(action['type'], 'ir.actions.act_window')
        self.assertEqual(action['res_model'], 'ksw.bank.file.export.wizard')
        self.assertEqual(action['target'], 'new')
        self.assertEqual(
            action['context']['default_payslip_run_id'], self.batch.id,
        )

    # ================================================================
    # Tests — Bank label uniqueness
    # ================================================================

    def test_bank_label_uses_acc_number(self):
        """_bank_label prefers acc_number over bank_id.name."""
        wiz = self._make_wizard()
        label_1 = wiz._bank_label(self.kawthar_bank_1)
        label_2 = wiz._bank_label(self.kawthar_bank_2)
        # Both have the same bank_id.name ('Al Rajhi Bank'),
        # but acc_number differs
        self.assertNotEqual(label_1, label_2)
        self.assertIn('Kawther', label_1)
        self.assertIn('Hayat', label_2)

    def test_bank_label_safe_for_filenames(self):
        """_bank_label replaces spaces with underscores."""
        wiz = self._make_wizard()
        label = wiz._bank_label(self.kawthar_bank_1)
        self.assertNotIn(' ', label)

    # ================================================================
    # Tests — WPS scoping (no Kawthar employees in WPS files)
    # ================================================================

    def test_wps_excel_excludes_kawthar_employees(self):
        """WPS Excel export only includes employees assigned to WPS bank."""
        wiz = self._make_wizard(mode='specific_excel', bank=self.wps_bank)
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        data = base64.b64decode(att.datas)
        wb = openpyxl.load_workbook(io.BytesIO(data))

        # Summary sheet should only have the WPS employee
        ws_summary = wb['Payroll Summary']
        names = []
        for r in range(2, ws_summary.max_row + 1):
            name = ws_summary.cell(r, 1).value
            if name:
                names.append(name)
        self.assertIn('AHMED WPS EMPLOYEE', names)
        self.assertNotIn('ANITA KAWTHAR ONE', names)
        self.assertNotIn('OLIVER KAWTHAR TWO', names)

    # ================================================================
    # Tests — Export order by employee x_payslip_export_order
    # ================================================================

    def test_specific_wps_txt_respects_export_order(self):
        """WPS TXT detail lines follow employee export order; unset is last."""
        emp_1 = self._create_employee(
            'ZZ ORDER ONE', '301', '3003003001',
            self.bank_anb, 'SA11111111111111111W0301',
        )
        emp_2 = self._create_employee(
            'AA ORDER TWO', '302', '3003003002',
            self.bank_anb, 'SA11111111111111111W0302',
        )
        emp_1.sudo().write({
            'x_salary_bank_account_id': self.wps_bank.id,
            'x_payslip_export_order': 1,
        })
        emp_2.sudo().write({
            'x_salary_bank_account_id': self.wps_bank.id,
            'x_payslip_export_order': 2,
        })

        self._create_payslip(
            emp_1, self.batch,
            basic=5000, hra=1000, gross=7000, deductions=200, net=6800,
        )
        self._create_payslip(
            emp_2, self.batch,
            basic=5000, hra=1000, gross=7000, deductions=150, net=6850,
        )

        wiz = self._make_wizard(
            mode='specific_txt', bank=self.wps_bank,
            value_date=date(2026, 3, 30),
        )
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)

        lines = [l for l in base64.b64decode(att.datas).decode('utf-8').split('\n') if l]
        detail_barcodes = [line[:12] for line in lines[1:]]
        self.assertEqual(
            detail_barcodes,
            ['000000000301', '000000000302', '000000000101'],
        )

    def test_specific_wps_excel_respects_export_order(self):
        """WPS Excel summary rows follow employee export order; unset is last."""
        emp_1 = self._create_employee(
            'ZZ EXCEL ONE', '401', '4004004001',
            self.bank_anb, 'SA11111111111111111W0401',
        )
        emp_2 = self._create_employee(
            'AA EXCEL TWO', '402', '4004004002',
            self.bank_anb, 'SA11111111111111111W0402',
        )
        emp_1.sudo().write({
            'x_salary_bank_account_id': self.wps_bank.id,
            'x_payslip_export_order': 1,
        })
        emp_2.sudo().write({
            'x_salary_bank_account_id': self.wps_bank.id,
            'x_payslip_export_order': 2,
        })

        self._create_payslip(
            emp_1, self.batch,
            basic=5000, hra=1000, gross=7000, deductions=200, net=6800,
        )
        self._create_payslip(
            emp_2, self.batch,
            basic=5000, hra=1000, gross=7000, deductions=150, net=6850,
        )

        wiz = self._make_wizard(mode='specific_excel', bank=self.wps_bank)
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)

        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(att.datas)))
        ws_summary = wb['Payroll Summary']
        self.assertEqual(
            self._data_block_names(ws_summary, 1, 2),
            ['ZZ EXCEL ONE', 'AA EXCEL TWO', 'AHMED WPS EMPLOYEE'],
        )

    def test_kawthar_excel_has_payroll_summary_sheet(self):
        """Kawthar Excel export includes both Payroll Summary and PREFORMAT PAYMENTS sheets."""
        wiz = self._make_wizard(mode='specific_excel', bank=self.kawthar_bank_1)
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(att.datas)))
        self.assertIn('Payroll Summary', wb.sheetnames)
        self.assertIn('PREFORMAT PAYMENTS', wb.sheetnames)
        # Payroll Summary must have the Kawthar employee name
        ws_summary = wb['Payroll Summary']
        names = [
            ws_summary.cell(r, 1).value
            for r in range(2, ws_summary.max_row + 1)
            if ws_summary.cell(r, 1).value
        ]
        self.assertIn('ANITA KAWTHAR ONE', names)
        self.assertNotIn('AHMED WPS EMPLOYEE', names)

    # ================================================================
    # Tests — All Excel mode
    # ================================================================

    def test_all_excel_produces_zip_for_multiple_banks(self):
        """all_excel with 3 banks → zip file with 3 Excel files."""
        wiz = self._make_wizard(mode='all_excel')
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        self.assertEqual(att.mimetype, 'application/zip')
        data = base64.b64decode(att.datas)
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()
        # 1 WPS + 2 Kawthar = 3 files
        self.assertEqual(len(names), 3)
        wps_files = [n for n in names if n.startswith('WPS_')]
        kaw_files = [n for n in names if n.startswith('Kawthar_')]
        self.assertEqual(len(wps_files), 1)
        self.assertEqual(len(kaw_files), 2)

    def test_all_excel_kawthar_filenames_unique(self):
        """Kawthar Excel filenames include unique acc_number."""
        wiz = self._make_wizard(mode='all_excel')
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        data = base64.b64decode(att.datas)
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            kaw_files = [n for n in zf.namelist() if n.startswith('Kawthar_')]
        self.assertEqual(len(kaw_files), 2)
        # Must be distinguishable
        self.assertNotEqual(kaw_files[0], kaw_files[1])

    # ================================================================
    # Tests — All TXT mode
    # ================================================================

    def test_all_txt_produces_zip(self):
        """all_txt → zip file with WPS + Kawthar text files."""
        wiz = self._make_wizard(mode='all_txt',
                                value_date=date(2026, 3, 30))
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        self.assertEqual(att.mimetype, 'application/zip')
        data = base64.b64decode(att.datas)
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()
        self.assertEqual(len(names), 3)  # 1 WPS + 2 Kawthar
        txt_files = [n for n in names if n.endswith('.txt')]
        self.assertEqual(len(txt_files), 3)

    # ================================================================
    # Tests — Specific bank Excel
    # ================================================================

    def test_specific_excel_single_file(self):
        """specific_excel with one bank → single file, not zip."""
        wiz = self._make_wizard(mode='specific_excel', bank=self.wps_bank)
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        # Not a zip
        self.assertNotEqual(att.mimetype, 'application/zip')
        self.assertTrue(att.name.endswith('.xlsx'))

    def test_specific_excel_requires_bank(self):
        """specific_excel without bank_account_id raises UserError."""
        wiz = self._make_wizard(mode='specific_excel')
        with self.assertRaises(UserError):
            wiz.action_export()

    # ================================================================
    # Tests — Specific bank TXT
    # ================================================================

    def test_specific_txt_single_file(self):
        """specific_txt → single text file download."""
        wiz = self._make_wizard(mode='specific_txt', bank=self.wps_bank,
                                value_date=date(2026, 3, 30))
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        self.assertTrue(att.name.endswith('.txt'))

    def test_specific_txt_requires_value_date(self):
        """specific_txt without value_date raises UserError."""
        wiz = self._make_wizard(mode='specific_txt', bank=self.wps_bank)
        wiz.value_date = False
        with self.assertRaises(UserError):
            wiz.action_export()

    # ================================================================
    # Tests — Kawthar TXT format
    # ================================================================

    def test_kawthar_txt_line_length(self):
        """Each Kawthar TXT line is exactly 194 characters."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        lines = [l for l in data.decode('utf-8').split('\n') if l]
        for line in lines:
            self.assertEqual(
                len(line), self.KAWTHAR_LINE_LEN,
                'Kawthar TXT line length mismatch: %d' % len(line),
            )

    def test_kawthar_txt_employee_id_is_row_sequence(self):
        """First 12 chars are the row's 1..N position in this file."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        line = data.decode('utf-8').split('\n')[0]
        self.assertEqual(line[:12], '000000000001')

    def test_kawthar_txt_cic(self):
        """Chars 13-22 are the CIC number zero-padded to 10."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        line = data.decode('utf-8').split('\n')[0]
        self.assertEqual(line[12:22], '0005555555')

    def test_kawthar_txt_card_number(self):
        """Chars 23-36 are the card number (acc_number) left-justified 14 chars."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        line = data.decode('utf-8').split('\n')[0]
        card = line[22:36]
        self.assertEqual(len(card), 14)
        self.assertTrue(card.startswith('56891100001302'))

    def test_kawthar_txt_employee_name(self):
        """Chars 37-86 are the employee name left-justified 50 chars."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        line = data.decode('utf-8').split('\n')[0]
        name = line[36:86]
        self.assertEqual(len(name), 50)
        self.assertTrue(name.startswith('ANITA KAWTHAR ONE'))

    def test_kawthar_txt_net_in_halalas(self):
        """Net amount field (15 chars) is in halalas."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        line = data.decode('utf-8').split('\n')[0]
        # national_id = 10 chars at pos 86-95
        # net = 15 chars at pos 96-110
        net_field = line[96:111]
        # 6500 SAR = 650000 halalas
        self.assertEqual(net_field, '000000000650000')

    def test_kawthar_txt_value_date(self):
        """Value date appears as YYYYMMDD at pos 111-118."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        line = data.decode('utf-8').split('\n')[0]
        self.assertEqual(line[111:119], '20260330')

    def test_kawthar_txt_operation_code(self):
        """Operation code is 1 char at pos 119."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30),
                                operation_code='2')
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        line = data.decode('utf-8').split('\n')[0]
        self.assertEqual(line[119], '2')

    def test_kawthar_txt_salary_breakdown_halalas(self):
        """Basic/HRA/Other/Deductions fields are in halalas."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        line = data.decode('utf-8').split('\n')[0]
        # After op_code(1) + zeros(6) + spaces(20) = 27 chars after pos 119
        # basic starts at pos 146
        basic = int(line[146:158])
        hra = int(line[158:170])
        other = int(line[170:182])
        ded = int(line[182:194])
        # 5000 SAR basic = 500000 halalas
        self.assertEqual(basic, 500000)
        # 1000 SAR hra = 100000 halalas
        self.assertEqual(hra, 100000)
        # other = gross - basic - hra = 7000 - 5000 - 1000 = 1000 SAR = 100000
        self.assertEqual(other, 100000)
        # 500 SAR deductions = 50000 halalas
        self.assertEqual(ded, 50000)

    def test_kawthar_txt_no_header_row(self):
        """Kawthar TXT has no header row — all lines are data."""
        wiz = self._make_wizard(mode='specific_txt',
                                bank=self.kawthar_bank_1,
                                value_date=date(2026, 3, 30))
        data = wiz._make_kawthar_txt(self.kawthar_bank_1,
                                     self.slip_kaw_1)
        lines = [l for l in data.decode('utf-8').split('\n') if l]
        # 1 employee → 1 line
        self.assertEqual(len(lines), 1)

    # ================================================================
    # Tests — Payroll summary sheet scoped to slips
    # ================================================================

    def test_summary_sheet_scoped_to_slips(self):
        """_fill_payroll_summary_sheet only includes the passed slips."""
        wb = openpyxl.Workbook()
        # Only pass the WPS slip
        self.batch._fill_payroll_summary_sheet(wb, self.slip_wps)
        ws = wb.active
        names = self._data_block_names(ws, 1, 2)
        self.assertEqual(len(names), 1)
        self.assertEqual(names[0], 'AHMED WPS EMPLOYEE')

    def test_summary_sheet_all_slips_default(self):
        """Without explicit slips, summary includes all batch slips."""
        wb = openpyxl.Workbook()
        self.batch._fill_payroll_summary_sheet(wb)
        ws = wb.active
        self.assertEqual(len(self._data_block_names(ws, 1, 2)), 3)

    # ================================================================
    # Tests — Error handling
    # ================================================================

    def test_error_empty_batch(self):
        """Wizard raises on empty batch."""
        empty_batch = self.env['hr.payslip.run'].create({
            'name': 'Empty',
            'date_start': date(2026, 3, 1),
            'date_end': date(2026, 3, 31),
        })
        wiz = self._make_wizard(batch=empty_batch, mode='all_excel')
        with self.assertRaises(UserError):
            wiz.action_export()

    def test_error_no_file_type(self):
        """Wizard raises when bank account has no x_file_type."""
        # Create a bank with no file type
        no_type_bank = self.env['res.partner.bank'].create({
            'acc_number': 'No Type Bank',
            'partner_id': self.company_partner.id,
            'bank_id': self.bank_anb.id,
        })
        self.emp_wps.sudo().write({
            'x_salary_bank_account_id': no_type_bank.id,
        })
        wiz = self._make_wizard(mode='all_excel')
        with self.assertRaises(UserError):
            wiz.action_export()
        # Restore
        self.emp_wps.sudo().write({
            'x_salary_bank_account_id': self.wps_bank.id,
        })

    def test_error_specific_no_slips_for_bank(self):
        """specific_excel raises when no payslips assigned to bank."""
        unused_bank = self.env['res.partner.bank'].create({
            'acc_number': 'Unused Bank',
            'partner_id': self.company_partner.id,
            'bank_id': self.bank_anb.id,
            'x_file_type': 'wps',
        })
        wiz = self._make_wizard(mode='specific_excel', bank=unused_bank)
        with self.assertRaises(UserError):
            wiz.action_export()

    # ================================================================
    # Tests — Excel shows every record, colour-coded by TXT inclusion
    #
    # The bank TXT keeps dropping unpayable rows; the Excel must show
    # them so the accountant can see WHY someone is missing (zero net
    # from a loan / ATT_DED, or a data problem).
    # ================================================================

    RED = 'FFC7CE'
    AMBER = 'FFEB9C'

    def _fill_rgb(self, cell):
        """Return the 6-char fill colour of a cell, or '' when unfilled.

        An unfilled openpyxl cell still reports rgb '00000000', so the
        pattern type is what actually distinguishes "no fill".
        """
        if not cell.fill.patternType:
            return ''
        rgb = cell.fill.fgColor.rgb
        return rgb[-6:] if isinstance(rgb, str) else ''

    def _sheet_rows(self, ws, name_col, first_row):
        """Return {employee name: row index} for a sheet's data rows."""
        return {
            ws.cell(r, name_col).value: r
            for r in range(first_row, ws.max_row + 1)
            if ws.cell(r, name_col).value
        }

    def _data_block_names(self, ws, name_col, first_row):
        """Names in the contiguous data block, stopping at the first blank.

        Everything past that blank is a labelled review block (banner,
        excluded rows, totals), not employee data.
        """
        names = []
        for r in range(first_row, ws.max_row + 1):
            v = ws.cell(r, name_col).value
            if not v:
                break
            names.append(v)
        return names

    def _export_wb(self, mode, bank):
        """Run an Excel export and return the loaded workbook."""
        wiz = self._make_wizard(mode=mode, bank=bank)
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        return openpyxl.load_workbook(io.BytesIO(base64.b64decode(att.datas)))

    def _export_txt(self, bank):
        """Run a TXT export and return the decoded content."""
        wiz = self._make_wizard(mode='specific_txt', bank=bank)
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        att = self.env['ir.attachment'].browse(att_id)
        return base64.b64decode(att.datas).decode('utf-8')

    def _add_wps_employee(self, name, barcode, id_number, net, iban='SA99'):
        """Create a WPS employee + payslip with an explicit NET."""
        emp = self._create_employee(name, barcode, id_number,
                                    self.bank_anb, iban)
        emp.sudo().write({'x_salary_bank_account_id': self.wps_bank.id})
        self._create_payslip(
            emp, self.batch,
            basic=5000, hra=1000, gross=6000,
            deductions=6000 - net, net=net,
        )
        return emp

    def test_zero_net_row_present_and_red_in_wps_excel(self):
        """A zero-NET employee is written to the WPS Excel, filled red."""
        self._add_wps_employee('ZERO NET GUY', '901', '9009009001', net=0)
        wb = self._export_wb('specific_excel', self.wps_bank)

        ws = wb['WPS Bank File']
        rows = self._sheet_rows(ws, 3, 8)          # col C = Employee Name
        self.assertIn('ZERO NET GUY', rows,
                      'zero-NET employee must still appear in the Excel')
        r = rows['ZERO NET GUY']
        self.assertEqual(self._fill_rgb(ws.cell(r, 1)), self.RED)
        self.assertIn('Zero net', ws.cell(r, 15).value)
        # The deduction that ate the salary must be visible (col J)
        self.assertEqual(ws.cell(r, 10).value, 6000)

        # A normally-paid employee stays unfilled
        paid = rows['AHMED WPS EMPLOYEE']
        self.assertEqual(self._fill_rgb(ws.cell(paid, 1)), '')

    def test_zero_net_row_red_in_payroll_summary(self):
        """The same colour coding applies to the Payroll Summary sheet."""
        self._add_wps_employee('ZERO NET GUY', '901', '9009009001', net=0)
        wb = self._export_wb('specific_excel', self.wps_bank)

        ws = wb['Payroll Summary']
        rows = self._sheet_rows(ws, 1, 2)
        r = rows['ZERO NET GUY']
        self.assertEqual(self._fill_rgb(ws.cell(r, 1)), self.RED)
        self.assertIn('Zero net', ws.cell(r, 18).value)
        self.assertEqual(
            self._fill_rgb(ws.cell(rows['AHMED WPS EMPLOYEE'], 1)), '')

    def test_zero_net_still_excluded_from_wps_txt(self):
        """TXT behaviour is unchanged, and its control totals still match."""
        self._add_wps_employee('ZERO NET GUY', '901', '9009009001', net=0)
        content = self._export_txt(self.wps_bank)

        self.assertNotIn('ZERO NET GUY', content)
        lines = content.strip('\n').split('\n')
        header, details = lines[0], lines[1:]
        # Header record count (10N at offset 42) must match the detail lines
        self.assertEqual(int(header[42:52]), len(details))
        # Header total is whole SAR (13N at offset 29); the detail NET is in
        # halalas (15N at offset 109, after ref/swift/filler/iban/name).
        detail_total = sum(int(d[109:124]) for d in details) / 100.0
        self.assertEqual(int(header[29:42]), int(round(detail_total)))

    def test_negative_net_red_in_excel_and_absent_from_txt(self):
        """Negative NET is treated exactly like zero: red, and not in the TXT.

        This previously diverged — WPS Excel wrote the row, WPS TXT dropped it.
        """
        self._add_wps_employee('OVER DEDUCTED', '902', '9009009002', net=-250)

        wb = self._export_wb('specific_excel', self.wps_bank)
        ws = wb['WPS Bank File']
        r = self._sheet_rows(ws, 3, 8)['OVER DEDUCTED']
        self.assertEqual(self._fill_rgb(ws.cell(r, 1)), self.RED)
        self.assertIn('Negative net', ws.cell(r, 15).value)

        self.assertNotIn('OVER DEDUCTED', self._export_txt(self.wps_bank))

    def test_kawthar_no_bank_account_amber_and_absent_from_txt(self):
        """A Kawthar employee with no payroll card is amber, not red."""
        emp = self._create_employee(
            'NO CARD EMPLOYEE', '903', '9009009003', self.bank_rajhi, None,
        )
        emp.sudo().write({
            'x_salary_bank_account_id': self.kawthar_bank_1.id,
        })
        self._create_payslip(
            emp, self.batch,
            basic=4000, hra=0, gross=4000, deductions=0, net=4000,
        )

        wb = self._export_wb('specific_excel', self.kawthar_bank_1)
        ws = wb['PREFORMAT PAYMENTS']
        r = self._sheet_rows(ws, 3, 5)['NO CARD EMPLOYEE']
        self.assertEqual(self._fill_rgb(ws.cell(r, 1)), self.AMBER)
        self.assertIn('No payroll card', ws.cell(r, 15).value)

        wiz = self._make_wizard(mode='specific_txt', bank=self.kawthar_bank_1)
        result = wiz.action_export()
        att_id = int(result['url'].split('/web/content/')[1].split('?')[0])
        content = base64.b64decode(
            self.env['ir.attachment'].browse(att_id).datas).decode('utf-8')
        self.assertNotIn('NO CARD EMPLOYEE', content)

    def test_kawthar_sequence_covers_payable_rows_only(self):
        """Column A stays a clean 1..N; excluded rows get a blank sequence."""
        emp = self._create_employee(
            'ZERO CARD EMPLOYEE', '904', '9009009004',
            self.bank_rajhi, '5689110000130257801',
        )
        emp.sudo().write({
            'x_salary_bank_account_id': self.kawthar_bank_1.id,
        })
        self._create_payslip(
            emp, self.batch,
            basic=4000, hra=0, gross=4000, deductions=4000, net=0,
        )

        ws = self._export_wb(
            'specific_excel', self.kawthar_bank_1)['PREFORMAT PAYMENTS']
        rows = self._sheet_rows(ws, 3, 5)
        seqs = [
            ws.cell(r, 1).value for r in sorted(rows.values())
            if ws.cell(r, 1).value
        ]
        self.assertEqual(seqs, list(range(1, len(seqs) + 1)))
        # The excluded row carries no sequence number
        self.assertFalse(ws.cell(rows['ZERO CARD EMPLOYEE'], 1).value)

    def test_excluded_rows_sit_below_every_payable_row(self):
        """Excluded rows are a trailing block, so the top stays uploadable."""
        self._add_wps_employee('ZERO NET GUY', '901', '9009009001', net=0)
        self._add_wps_employee('PAID LATER', '905', '9009009005', net=3000)

        ws = self._export_wb('specific_excel', self.wps_bank)['WPS Bank File']
        rows = self._sheet_rows(ws, 3, 8)
        payable = [rows['AHMED WPS EMPLOYEE'], rows['PAID LATER']]
        self.assertGreater(rows['ZERO NET GUY'], max(payable))

    def test_generation_skips_on_summary_sheet_only(self):
        """Employees with no payslip are listed on the summary sheet only."""
        skipped_emp = self._create_employee(
            'NEVER GENERATED', '906', '9009009006', self.bank_anb, None,
        )
        self.env['ksw.payslip.run.skip.line'].create({
            'run_id': self.batch.id,
            'employee_id': skipped_emp.id,
            'reason': 'No active contract / salary structure for this period',
        })

        wb = self._export_wb('specific_excel', self.wps_bank)

        ws_summary = wb['Payroll Summary']
        rows = self._sheet_rows(ws_summary, 1, 2)
        self.assertIn('NEVER GENERATED', rows)
        r = rows['NEVER GENERATED']
        self.assertEqual(self._fill_rgb(ws_summary.cell(r, 1)), self.AMBER)
        self.assertIn('No active contract', ws_summary.cell(r, 18).value)

        # They have no resolved paying bank, so they must not reach the
        # bank sheet — it is an upload template.
        ws_bank = wb['WPS Bank File']
        self.assertNotIn('NEVER GENERATED', self._sheet_rows(ws_bank, 3, 8))

    # ================================================================
    # Tests — batch totals must state what the BANK pays
    # ================================================================

    def _totals_row(self, ws, prefix, money_col):
        """Return the amount on the totals line whose label starts prefix."""
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.startswith(prefix):
                return ws.cell(r, money_col).value
        return None

    def _wps_bank_total(self):
        self.batch._refresh_bank_totals()
        return self.batch.x_bank_total_ids.filtered(
            lambda t: t.bank_account_id == self.wps_bank
        )

    def test_bank_total_excludes_negative_net(self):
        """An over-deducted slip must not be netted off the payable total.

        Reproduces KSWCO batch 246: 4 slips at -390 dragged the displayed
        Kawthar total to 362,027 while the bank file paid 363,587.
        """
        self._add_wps_employee('OVER DEDUCTED', '910', '9009009010', net=-390)

        total = self._wps_bank_total()
        self.assertEqual(total.total_net, 11000,
                         'Total NET must be what the bank debits')
        self.assertEqual(total.excluded_net, -390)
        self.assertEqual(total.excluded_count, 1)
        self.assertEqual(total.payable_count, 1)
        self.assertEqual(total.slip_count, 2)

    def test_bank_total_excludes_zero_net(self):
        """Zero-NET rows inflate the employee count but pay nothing."""
        self._add_wps_employee('ZERO NET GUY', '911', '9009009011', net=0)

        total = self._wps_bank_total()
        self.assertEqual(total.total_net, 11000)
        self.assertEqual(total.payable_count, 1)
        self.assertEqual(total.excluded_count, 1)
        self.assertEqual(total.excluded_net, 0)
        self.assertEqual(total.slip_count, 2)

    def test_bank_total_matches_the_txt_control_total(self):
        """Total NET on the batch equals the WPS header's control total."""
        self._add_wps_employee('OVER DEDUCTED', '912', '9009009012', net=-390)
        self._add_wps_employee('ZERO NET GUY', '913', '9009009013', net=0)
        self._add_wps_employee('PAID TOO', '914', '9009009014', net=2500)

        total = self._wps_bank_total()
        header = self._export_txt(self.wps_bank).split('\n')[0]
        # Header layout: 12 filler + 'G' + 8 + 8, then the 13-digit total.
        self.assertEqual(int(header[29:42]), int(total.total_net))
        self.assertEqual(int(header[42:52]), total.payable_count)
        self.assertEqual(total.total_net, 13500)

    def test_excel_states_payable_and_batch_totals_separately(self):
        """Summing the Net column can no longer be mistaken for the payment."""
        self._add_wps_employee('OVER DEDUCTED', '915', '9009009015', net=-390)

        ws = self._export_wb('specific_excel', self.wps_bank)['Payroll Summary']
        self.assertEqual(
            self._totals_row(ws, 'Total written to the bank file', 15), 11000,
        )
        self.assertEqual(
            self._totals_row(ws, 'Excluded from the bank file', 15), -390,
        )
        self.assertEqual(
            self._totals_row(ws, 'Batch total', 15), 10610,
        )

    def test_bank_sheet_totals_carry_the_review_only_banner(self):
        """Totals on an upload sheet must be flagged as rows to delete."""
        self._add_wps_employee('OVER DEDUCTED', '916', '9009009016', net=-390)

        ws = self._export_wb('specific_excel', self.wps_bank)['WPS Bank File']
        self.assertEqual(
            self._totals_row(ws, 'Total written to the bank file', 6), 11000,
        )
        labels = [
            ws.cell(r, 1).value for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, str)
        ]
        self.assertTrue(
            any(l.startswith('Totals') and 'delete' in l for l in labels),
            'the totals block must warn that it is not part of the upload',
        )

