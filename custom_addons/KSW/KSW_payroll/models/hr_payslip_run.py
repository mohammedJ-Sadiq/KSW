from odoo import fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    openpyxl = None

# Excel row styling per export status: status -> (fill colour, font colour).
# 'ok' is deliberately absent — included rows carry no fill.
EXPORT_ROW_STYLES = {
    'excluded_zero': ('FFC7CE', '9C0006'),   # red   — NET <= 0, not payable
    'excluded_other': ('FFEB9C', '9C6500'),  # amber — dropped for another reason
    'warning': ('FFEB9C', '9C6500'),         # amber — in the file but needs a look
}

EXCLUDED_STATUSES = ('excluded_zero', 'excluded_other')


class KswPayslipRunBankTotal(models.Model):
    """Per-bank-account NET total summary for a payslip batch.

    Refreshed automatically after compute_sheet(), action_payslip_done(),
    and whenever the batch-level fallback bank account changes.
    """
    _name = 'ksw.payslip.run.bank.total'
    _description = 'Payslip Batch — Bank Account NET Total'
    _order = 'total_net desc'

    run_id = fields.Many2one(
        'hr.payslip.run', string='Payslip Batch',
        required=True, ondelete='cascade', index=True,
    )
    bank_account_id = fields.Many2one(
        'res.partner.bank', string='Bank Account',
        ondelete='set null',
    )
    bank_name = fields.Char(
        related='bank_account_id.bank_id.name',
        string='Bank Name',
        store=True,
    )
    file_type = fields.Selection(
        related='bank_account_id.x_file_type',
        string='Type',
        store=True,
    )
    slip_count = fields.Integer(
        string='Employees',
        help='Payslips in this batch resolved to this bank account, '
             'payable or not.',
    )
    payable_count = fields.Integer(
        string='In Bank File',
        help='Rows the bank text file actually carries.',
    )
    total_net = fields.Float(
        string='Total NET', digits=(16, 2),
        help='What the bank will debit: the sum of the rows written to the '
             'text file. Zero and negative NET rows are NOT netted off this '
             'figure — they are reported separately.',
    )
    excluded_count = fields.Integer(
        string='Excluded',
        help='Rows dropped from the bank text file (zero or negative NET, '
             'or no payroll card).',
    )
    excluded_net = fields.Float(
        string='Excluded NET', digits=(16, 2),
        help='Sum of NET on the dropped rows. Negative when an employee was '
             'over-deducted.',
    )


class KswPayslipRunSkipLine(models.Model):
    """Records employees the batch could not process normally.

    Two kinds of row, told apart by ``line_type`` — a skipped employee has
    no payslip at all, a warned one has a payslip that is very likely wrong.
    They must not be conflated: the export's "no payslip in this batch"
    section is built from the skipped rows only.
    """
    _name = 'ksw.payslip.run.skip.line'
    _description = 'Payslip Batch — Skipped Employee Log'
    _order = 'employee_id'

    run_id = fields.Many2one(
        'hr.payslip.run', string='Payslip Batch',
        required=True, ondelete='cascade', index=True,
    )
    employee_id = fields.Many2one(
        'hr.employee', string='Employee',
        required=True, ondelete='cascade',
    )
    reason = fields.Char(string='Reason', required=True)
    line_type = fields.Selection([
        ('skipped', 'Skipped — no payslip generated'),
        ('warning', 'Processed — figures need review'),
    ], string='Outcome', default='skipped', required=True)


class HrPayslipRun(models.Model):
    _inherit = 'hr.payslip.run'

    x_salary_bank_account_id = fields.Many2one(
        'res.partner.bank',
        string='Salary Paying Bank Account',
        help='Default company bank account for WPS export. '
             'Employees with their own Salary Paying Bank Account '
             'will override this.',
    )

    x_skip_line_ids = fields.One2many(
        'ksw.payslip.run.skip.line', 'run_id',
        string='Skipped Employees',
        help='Employees that were automatically skipped during payslip '
             'generation and the reason they were excluded.',
    )

    x_bank_total_ids = fields.One2many(
        'ksw.payslip.run.bank.total', 'run_id',
        string='Bank Account Totals',
        readonly=True,
    )

    def write(self, vals):
        res = super().write(vals)
        if 'x_salary_bank_account_id' in vals:
            self._refresh_bank_totals()
        return res

    def _refresh_bank_totals(self):
        """Recompute per-bank NET totals from the current slip_ids.

        ``total_net`` answers one question only: **what will the bank debit?**
        So it is built from the same classifier the export files use and sums
        the payable rows alone. Zero and negative NET rows are counted and
        summed separately instead of being netted off the payable figure.

        Summing every slip is what made this screen disagree with the file:
        on the KSWCO July 2026 batch, 4 over-deducted slips at -390 SAR pulled
        the Kawthar total down to 362,027 while the text file paid 363,587,
        and the 285 "Employees" included 44 zero-NET rows the bank never sees.

        Grouping is delegated to ``_group_slips_by_bank_account`` so the
        summary and the export wizard can never resolve banks differently.
        """
        BankTotal = self.env['ksw.payslip.run.bank.total'].sudo()
        for run in self:
            BankTotal.search([('run_id', '=', run.id)]).unlink()
            for bank, slips in run._group_slips_by_bank_account().items():
                payable_count = payable_net = 0
                excluded_count = excluded_net = 0
                classified = run._classify_export_slips(
                    slips, bank.x_file_type or 'wps',
                )
                for slip, status, _reason in classified:
                    net = run._get_line_total(slip, 'NET')
                    if status in EXCLUDED_STATUSES:
                        excluded_count += 1
                        excluded_net += net
                    else:
                        payable_count += 1
                        payable_net += net
                BankTotal.create({
                    'run_id': run.id,
                    'bank_account_id': bank.id or False,
                    'slip_count': len(slips),
                    'payable_count': payable_count,
                    'total_net': payable_net,
                    'excluded_count': excluded_count,
                    'excluded_net': excluded_net,
                })

    def draft_payslip_run(self):
        """Reopening a closed batch is a Payroll Manager action only.

        Same rationale as `hr.payslip._check_payroll_manager` — the stock
        button is open to every payroll Officer.
        """
        if not self.env.su and not self.env.user.has_group(
                'om_hr_payroll.group_hr_payroll_manager'):
            raise UserError(_(
                'Only a Payroll Manager may set a payslip batch back to '
                'draft. Please ask the payroll administrator to do it.'))
        return super().draft_payslip_run()

    def done_payslip_run(self):
        """Override: suppress per-slip bank total refresh during bulk validation.
        Calls _refresh_bank_totals() once per batch instead of once per slip.
        """
        for line in self.slip_ids:
            line.with_context(_ksw_skip_bank_refresh=True).action_payslip_done()
        self.write({'state': 'done'})
        self._refresh_bank_totals()
        return True

    def action_refresh_bank_totals(self):
        self.ensure_one()
        self._refresh_bank_totals()
        return True

    def action_clear_skip_log(self):
        """Remove all skipped-employee log entries for this batch."""
        self.ensure_one()
        self.x_skip_line_ids.unlink()
        return True


    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_line_total(self, slip, code):
        """Return the total of a salary rule line by code, or 0."""
        line = slip.line_ids.filtered(lambda l: l.code == code)
        return line[:1].total if line else 0.0

    def _get_wd_amount(self, slip, code):
        """Return the amount of a worked-day line by code, or 0."""
        wd = slip.worked_days_line_ids.filtered(lambda w: w.code == code)
        return wd[:1].amount if wd else 0.0

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------

    def action_open_export_wizard(self):
        """Open the unified bank file export wizard."""
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Export Bank File'),
            'res_model': 'ksw.bank.file.export.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_payslip_run_id': self.id,
            },
        }

    def action_open_batch_payslips(self):
        """Open batch payslip lines in the standard searchable payslip view."""
        self.ensure_one()
        action = self.env['ir.actions.actions']._for_xml_id(
            'om_hr_payroll.action_view_hr_payslip_form'
        )
        action.update({
            'name': _('Payslips - %s') % (self.name or _('Batch')),
            'domain': [('payslip_run_id', '=', self.id)],
            'context': {
                'default_payslip_run_id': self.id,
                'search_default_payslip_run_id': self.id,
            },
        })
        return action

    def _group_slips_by_bank_account(self):
        """Group payslips by the paying bank account

        Resolution order for each slip:
        1. Employee's ``x_salary_bank_account_id``
        2. Batch-level ``x_salary_bank_account_id`` (fallback)

        Returns a dict ``{res.partner.bank recordset: slip recordsets}``.
        Slips with **no** resolved bank account are collected under an
        empty recordset key so callers can report them.
        """
        groups = {}
        bank_model = self.env['res.partner.bank']
        for slip in self.slip_ids:
            bank = (
                slip.employee_id.sudo().x_salary_bank_account_id
                or self.x_salary_bank_account_id
                or bank_model
            )
            groups.setdefault(bank, self.env['hr.payslip'])
            groups[bank] |= slip
        return groups

    def _sorted_export_slips(self, slips):
        """Sort slips by configured employee export order, then name.

        ``x_payslip_export_order`` values <= 0 are treated as unset and sent
        to the end so explicitly configured employees always come first.
        """
        return slips.sorted(
            lambda s: (
                not bool(s.employee_id.sudo().x_payslip_export_order > 0),
                s.employee_id.sudo().x_payslip_export_order
                if s.employee_id.sudo().x_payslip_export_order > 0 else 0,
                s.employee_id.name or '',
                s.employee_id.id,
            )
        )

    # ------------------------------------------------------------------
    # Export classification — shared by every Excel sheet
    # ------------------------------------------------------------------

    def _classify_export_slips(self, slips, file_type='wps'):
        """Return ``[(slip, status, reason)]`` in ``_sorted_export_slips`` order.

        ``status`` mirrors exactly what the bank **text** file does with the row,
        so the Excel colour is never a guess:

        ``ok``
            written to the text file, no fill
        ``excluded_zero``
            NET <= 0, dropped from the text file (red)
        ``excluded_other``
            dropped from the text file for another reason (amber)
        ``warning``
            written to the text file but needs attention (amber)

        The text-file builders keep their own predicates; this must stay in
        agreement with them (``_build_wps_text``, ``_build_kawthar_text``).
        """
        rows = []
        for slip in self._sorted_export_slips(slips):
            net = self._get_line_total(slip, 'NET')
            bank = slip.employee_id.sudo().primary_bank_account_id
            if not net:
                status = 'excluded_zero'
                reason = _('Zero net salary — fully absorbed by deductions')
            elif net < 0:
                status = 'excluded_zero'
                reason = _('Negative net (%.2f) — over-deducted', net)
            elif not bank and file_type == 'kawthar':
                status = 'excluded_other'
                reason = _('No payroll card / bank account on the employee')
            elif not bank:
                status = 'warning'
                reason = _('No IBAN on the employee — still written to the '
                           'text file')
            else:
                status = 'ok'
                reason = _('Included in bank text file')
            rows.append((slip, status, reason))
        return rows

    def _style_export_row(self, ws, row_idx, ncols, status):
        """Apply the status fill/font to every cell of an Excel row."""
        style = EXPORT_ROW_STYLES.get(status)
        if not style:
            return
        fill_rgb, font_rgb = style
        fill = PatternFill('solid', fgColor=fill_rgb)
        font = Font(color=font_rgb)
        for ci in range(1, ncols + 1):
            c = ws.cell(row_idx, ci)
            c.fill = fill
            c.font = font

    def _write_export_banner(self, ws, last_row, text):
        """Write a bold separator banner below ``last_row``.

        Row ``last_row + 1`` is left blank so the banner is visually detached
        from the rows above it. Returns the first free row index after the
        banner.
        """
        c = ws.cell(last_row + 2, 1, text)
        c.font = Font(bold=True, size=11, color='9C6500')
        return last_row + 3

    def _write_export_totals(self, ws, last_row, money_col,
                             payable_count, payable_net,
                             excluded_count, excluded_net,
                             banner=False):
        """Write the payable / excluded / batch totals under a data block.

        The payable line is the only one that matches the bank text file, so
        it is stated separately from the batch total instead of leaving the
        reader to sum the money column — that sum silently nets the negative
        (over-deducted) rows off the amount the bank will actually debit.

        ``banner`` adds the review-only warning used on the bank upload
        sheets, where every row below the payable block must be deleted
        before uploading. Returns the index of the last row written.
        """
        bold = Font(bold=True, size=11)
        ri = last_row + 2
        if banner:
            c = ws.cell(ri, 1, _('Totals — review only, delete these rows '
                                 'before uploading'))
            c.font = Font(bold=True, size=11, color='9C6500')
            ri += 2
        rows = [(
            _('Total written to the bank file (%s rows)', payable_count),
            payable_net,
        )]
        if excluded_count:
            rows.append((
                _('Excluded from the bank file (%s rows)', excluded_count),
                excluded_net,
            ))
            rows.append((
                _('Batch total (all %s rows)',
                  payable_count + excluded_count),
                payable_net + excluded_net,
            ))
        for label, amount in rows:
            ws.cell(ri, 1, label).font = bold
            ws.cell(ri, money_col, amount).font = bold
            ri += 1
        return ri - 1

    # ------------------------------------------------------------------
    # Sheet 1 — Internal payroll summary
    # ------------------------------------------------------------------

    def _fill_payroll_summary_sheet(self, wb, slips=None, file_type='wps'):
        if slips is None:
            slips = self.slip_ids
        ws = wb.active
        ws.title = 'Payroll Summary'

        headers = [
            'Employee', 'SSN No', 'Date From', 'Date To', 'Department',
            'Basic Salary', 'House Rent Allowance', 'Other Allowance',
            'Gross', 'Absence Deduction', 'Attendance Deductions',
            'Missed Days (ATT SHEET)', 'Social Insurance', 'Loan',
            'Net Salary', 'Bank Account Number', 'Bank Name',
            'Bank File Status',
        ]

        hdr_font = Font(bold=True, size=11)
        hdr_fill = PatternFill('solid', fgColor='D9E1F2')
        hdr_align = Alignment(horizontal='center', wrap_text=True)
        thin = Border(
            left=Side('thin'), right=Side('thin'),
            top=Side('thin'), bottom=Side('thin'),
        )

        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = thin

        ri = 1
        payable_count = payable_net = 0
        excluded_count = excluded_net = 0
        for slip, status, reason in self._classify_export_slips(
            slips, file_type,
        ):
            ri += 1
            emp = slip.employee_id
            bank = emp.sudo().primary_bank_account_id
            is_sheet = emp.sudo().x_is_attendance_sheet

            basic = self._get_line_total(slip, 'BASIC')
            hra = self._get_line_total(slip, 'HRA')
            gross = self._get_line_total(slip, 'GROSS')
            gosi = self._get_line_total(slip, 'GOSI')
            net = self._get_line_total(slip, 'NET')
            attded = self._get_line_total(slip, 'ATTDED')

            # Other allowance = GROSS - BASIC - HRA
            other_alw = gross - basic - hra if gross else 0.0

            # Deductions breakdown
            abs_ded = 0.0
            att_ded = 0.0
            sheet_ded = 0.0
            if is_sheet:
                # Sheet employee: deduction goes to column L
                sheet_ded = -self._get_wd_amount(slip, 'ATT_DED')
            else:
                # Biometric: absence (J) and late/early (K)
                abs_ded = -self._get_wd_amount(slip, 'ATT_ABS')
                att_ded = -(
                    self._get_wd_amount(slip, 'ATT_LATE')
                    + self._get_wd_amount(slip, 'ATT_EARLY')
                )

            # Loan = total DED minus ATTDED and GOSI
            total_ded_cat = sum(
                l.total for l in slip.line_ids
                if l.category_id.code == 'DED'
            )
            loan = total_ded_cat - attded - gosi

            row = [
                emp.name or '',
                emp.ssnid or emp.identification_id or '',
                slip.date_from,
                slip.date_to,
                emp.department_id.name if emp.department_id else '',
                basic or None,
                hra or None,
                other_alw or None,
                gross or None,
                abs_ded or None,
                att_ded or None,
                sheet_ded or None,
                gosi or None,
                loan or None,
                net,
                bank.acc_number if bank else '',
                bank.bank_id.name if bank and bank.bank_id else '',
                reason,
            ]
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = thin
            self._style_export_row(ws, ri, len(headers), status)

            if status in EXCLUDED_STATUSES:
                excluded_count += 1
                excluded_net += net
            else:
                payable_count += 1
                payable_net += net

        # 'Net Salary' is column 15 — see `headers` above.
        ri = self._write_export_totals(
            ws, ri, 15, payable_count, payable_net,
            excluded_count, excluded_net,
        )

        # Employees that never got a payslip in this batch. They have no
        # resolved paying bank, so they cannot be attributed to a single bank
        # file — they belong on this internal sheet only. Warning rows are
        # excluded: those employees DO have a payslip, listed above.
        skipped_lines = self.x_skip_line_ids.filtered(
            lambda l: l.line_type != 'warning')
        if skipped_lines:
            ri = self._write_export_banner(
                ws, ri, _('Employees with no payslip in this batch'),
            )
            for line in skipped_lines:
                ws.cell(ri, 1, line.employee_id.name or '').border = thin
                ws.cell(ri, len(headers), line.reason or '').border = thin
                self._style_export_row(
                    ws, ri, len(headers), 'excluded_other',
                )
                ri += 1

        # Auto-width columns
        for ci in range(1, len(headers) + 1):
            letter = openpyxl.utils.get_column_letter(ci)
            mx = max(
                len(str(ws.cell(row=r, column=ci).value or ''))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[letter].width = min(mx + 3, 35)

    # ------------------------------------------------------------------
    # Sheet 2 — WPS bank upload (Al Rajhi format)
    # ------------------------------------------------------------------

    def _fill_wps_sheet(self, wb, bank_account, slips=None, suffix=''):
        if slips is None:
            slips = self.slip_ids
        sheet_name = ('WPS Bank File%s' % suffix)[:31]  # Excel 31-char limit
        ws = wb.create_sheet(sheet_name)

        bold = Font(bold=True, size=11)
        title_font = Font(bold=True, size=12, color='003366')
        thin = Border(
            left=Side('thin'), right=Side('thin'),
            top=Side('thin'), bottom=Side('thin'),
        )
        hdr_fill = PatternFill('solid', fgColor='C6EFCE')
        ar_fill = PatternFill('solid', fgColor='E2EFDA')

        # Read WPS header values from the selected bank account
        cic = bank_account.x_wps_cic_number or ''
        debit = bank_account.x_wps_debit_account or ''
        mol = bank_account.x_wps_mol_id or ''

        # ── Row 1 ──
        ws.cell(1, 1, 'CIC - رقم العميل').font = bold
        ws.cell(1, 2, cic)
        ws.merge_cells('C1:D1')
        ws.cell(1, 3,
                'Alrajhi Bank WPS Payroll Payments Upload File'
                ).font = title_font

        # ── Row 2 ──
        ws.cell(2, 1, 'Debit Account:').font = bold
        ws.cell(2, 2, debit)
        ws.merge_cells('C2:D2')
        ws.cell(2, 3,
                'Notes: Template used for upload of WPS Payroll data'
                ).font = Font(bold=True, size=10)
        ws.cell(2, 8, 'Type of Payroll')
        ws.cell(2, 9, 'WPS')

        # ── Row 3 ──
        ws.cell(3, 1, 'MOL ID').font = bold
        ws.cell(3, 2, mol)

        # ── Row 4 ──
        ws.cell(4, 1, 'Payment Purpose').font = bold
        ws.cell(4, 2, 'Payroll')

        # ── Row 5 ──
        ws.cell(5, 1, 'Company Remarks').font = bold
        ws.cell(5, 2, 'Payroll')

        # ── Row 6–7: English / Arabic headers ──
        en_headers = [
            'Bank Name', 'Account Number(34N)', 'Employee Name',
            'Employee Number', 'National ID Number', 'Salary (15N)',
            'Basic Salary', 'Housing Allowance', 'Other Earnings',
            'Deductions', 'Branch Code', 'Branch Name',
            'Employee Remarks', 'Employee Department',
            # Column O — past the 14-column bank template, review use only
            'Status',
        ]
        ar_headers = [
            'بنك الموظف', 'رقم أيبان الموظف', 'إسم الموظف',
            'الرقم الوظيفي', 'رقم الهوية للموظف', 'إجمالي الراتب',
            'الراتب الأساسي', 'بدل السكن', 'بدل أخرى',
            'الخصومات', 'رمز الفرع', 'اسم الفرع',
            'ملاحظات الموظف', 'قسم الموظف',
            'الحالة',
        ]

        for ci, (e, a) in enumerate(zip(en_headers, ar_headers), 1):
            c6 = ws.cell(6, ci, e)
            c6.font = bold
            c6.fill = hdr_fill
            c6.border = thin
            c6.alignment = Alignment(horizontal='center')

            c7 = ws.cell(7, ci, a)
            c7.font = Font(bold=True, size=10)
            c7.fill = ar_fill
            c7.border = thin
            c7.alignment = Alignment(horizontal='center')

        # ── Data rows (row 8+) ──
        # Every payslip is written, so the reviewer can see WHY someone was
        # left out of the text file. Rows the text file drops are pushed into
        # a trailing block behind a banner, keeping the top of the sheet
        # upload-ready.
        classified = self._classify_export_slips(slips, 'wps')
        included = [r for r in classified if r[1] not in EXCLUDED_STATUSES]
        excluded = [r for r in classified if r[1] in EXCLUDED_STATUSES]

        def _write(row_idx, slip, status, reason):
            emp = slip.employee_id
            bank = emp.sudo().primary_bank_account_id

            net = self._get_line_total(slip, 'NET')
            basic = self._get_line_total(slip, 'BASIC')
            hra = self._get_line_total(slip, 'HRA')
            gross = self._get_line_total(slip, 'GROSS')
            other_e = gross - basic - hra if gross else 0.0

            total_ded = abs(sum(
                l.total for l in slip.line_ids
                if l.category_id.code == 'DED'
            )) or 0.0

            data = [
                bank.bank_id.name if bank and bank.bank_id else '',
                bank.acc_number if bank else '',
                emp.name or '',
                # Must be the same identifier the WPS TXT writes in its first
                # 12 chars (x_employee_no). barcode is the biometric device ID
                # and is empty for every WPS employee in KSWCO, which left this
                # column blank in the Excel while the TXT carried a number.
                emp.sudo().x_employee_no or emp.barcode or '',
                emp.ssnid or emp.identification_id or '',
                net,              # F: Salary (net)
                basic,            # G: Basic
                hra,              # H: Housing
                other_e,          # I: Other earnings
                total_ded,        # J: Deductions
                '',               # K: Branch Code
                '',               # L: Branch Name
                '',               # M: Employee Remarks
                emp.department_id.name if emp.department_id else '',
                reason,           # O: Status (review only)
            ]
            for ci, v in enumerate(data, 1):
                c = ws.cell(row_idx, ci, v)
                c.border = thin
            self._style_export_row(ws, row_idx, len(en_headers), status)

        ri = 8
        for slip, status, reason in included:
            _write(ri, slip, status, reason)
            ri += 1

        if excluded:
            ri = self._write_export_banner(ws, ri - 1, _(
                'Excluded from the bank text file — review only, '
                'delete these rows before uploading'
            ))
            for slip, status, reason in excluded:
                _write(ri, slip, status, reason)
                ri += 1

        # 'Salary (15N)' is column 6 — see `en_headers` above.
        self._write_export_totals(
            ws, ri - 1, 6,
            len(included), sum(self._get_line_total(r[0], 'NET')
                               for r in included),
            len(excluded), sum(self._get_line_total(r[0], 'NET')
                               for r in excluded),
            banner=True,
        )

        # Auto-width columns
        for ci in range(1, len(en_headers) + 1):
            letter = openpyxl.utils.get_column_letter(ci)
            mx = max(
                len(str(ws.cell(row=r, column=ci).value or ''))
                for r in range(1, max(ws.max_row + 1, 2))
            )
            ws.column_dimensions[letter].width = min(mx + 3, 40)

