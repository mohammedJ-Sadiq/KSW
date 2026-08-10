"""Bank-file export wizard for the monthly commission payment.

Generates WPS Excel / Kawthar TXT files from ``net_payable`` on each
``ksw.pay.run.line`` — the payment register built when the General Manager
approves the month. Same file formats as ``KSW_payroll``; different source.
"""
import base64
import io
import zipfile

from odoo import _, api, fields, models
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None


EXPORT_MODES = [
    ('all_excel',    'All banks – Excel files'),
    ('all_txt',      'All banks – Text files (Kawthar)'),
    ('specific_excel', 'Specific bank – Excel'),
    ('specific_txt', 'Specific bank – Text file'),
]


class KswCommissionBankExportWizard(models.TransientModel):
    _name = 'ksw.commission.bank.export.wizard'
    _description = 'KSW Commission Bank File Export Wizard'

    run_id = fields.Many2one(
        'ksw.pay.run', required=True, readonly=True,
    )
    export_mode = fields.Selection(
        EXPORT_MODES, required=True, default='all_excel',
    )
    bank_account_id = fields.Many2one(
        'res.partner.bank',
        domain="[('x_file_type', '!=', False),"
               " ('partner_id', '=', company_partner_id)]",
    )
    company_partner_id = fields.Many2one(
        'res.partner', compute='_compute_company_partner',
    )
    value_date = fields.Date(
        default=fields.Date.context_today,
        help='Payment value date used in TXT files.',
    )
    operation_code = fields.Selection(
        [('1', '1 – New'), ('2', '2 – Renewal'), ('3', '3 – Delete')],
        default='2', string='Kawthar Operation',
    )

    @api.depends('run_id')
    def _compute_company_partner(self):
        for rec in self:
            rec.company_partner_id = rec.env.company.partner_id

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _group_and_validate(self, require_type=None):
        run = self.run_id
        if not run.line_ids:
            raise UserError(_('The payment register is empty — there is nothing to export.'))
        groups = run._group_lines_by_bank_account()
        no_bank = groups.pop(self.env['res.partner.bank'], None)
        if no_bank:
            names = ', '.join(no_bank.mapped('employee_id.name'))
            raise UserError(_(
                'The following employees have no bank account and no '
                'run-level fallback:\n%s', names))
        no_type = [b for b in groups if not b.x_file_type]
        if no_type:
            accs = ', '.join(b.acc_number or str(b.id) for b in no_type)
            raise UserError(_(
                'These bank accounts have no Payroll File Type set:\n%s', accs))
        if require_type:
            groups = {b: s for b, s in groups.items() if b.x_file_type == require_type}
        return groups

    def _batch_label(self):
        return (self.run_id.name or '').replace(' ', '_').replace('/', '-')

    def _bank_label(self, bank):
        return (bank.acc_number or bank.bank_id.name or str(bank.id)
                ).replace(' ', '_').replace('/', '-')

    def _bundle_and_download(self, files):
        if not files:
            raise UserError(_('No files were generated.'))
        run = self.run_id
        if len(files) == 1:
            fname, data = files[0]
            mimetype = (
                'text/plain' if fname.endswith('.txt')
                else 'application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet'
            )
            att = self.env['ir.attachment'].create({
                'name': fname, 'type': 'binary',
                'datas': base64.b64encode(data),
                'mimetype': mimetype,
                'res_model': run._name, 'res_id': run.id,
            })
        else:
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for fname, data in files:
                    zf.writestr(fname, data)
            att = self.env['ir.attachment'].create({
                'name': 'CommissionsBank_%s.zip' % self._batch_label(),
                'type': 'binary',
                'datas': base64.b64encode(buf.getvalue()),
                'mimetype': 'application/zip',
                'res_model': run._name, 'res_id': run.id,
            })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % att.id,
            'target': 'new',
        }

    # --- Excel generation --------------------------------------------------

    def _make_comm_summary_excel(self, wb, lines):
        """Fill a summary worksheet (mirrors payroll summary but for commissions)."""
        ws = wb.active
        ws.title = 'Commission Summary'
        thin = Border(
            left=Side('thin'), right=Side('thin'),
            top=Side('thin'), bottom=Side('thin'),
        )
        bold = Font(bold=True, size=11)
        hdr_fill = PatternFill('solid', fgColor='D9E1F2')
        headers = [
            'Employee', 'SSN', 'Department',
            'Manual Lines', 'Entry Sheets', 'Gross Total',
            'Loans Deduction', 'Bank Transfer Amount',
            'Bank Account', 'Bank Name',
        ]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = bold
            c.fill = hdr_fill
            c.border = thin
            c.alignment = Alignment(horizontal='center', wrap_text=True)
        for ri, line in enumerate(
                lines.sorted(lambda s: s.employee_id.name or ''), 2):
            emp = line.employee_id.sudo()
            bank = getattr(emp, 'x_salary_bank_account_id', False)
            row = [
                emp.name or '',
                emp.identification_id or '',
                emp.department_id.name if emp.department_id else '',
                line.earnings,
                line.loan_offset,
                line.net_payable,
                bank.acc_number if bank else '',
                bank.bank_id.name if bank and bank.bank_id else '',
            ]
            for ci, v in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.border = thin

    def _make_wps_excel(self, bank, lines):
        if not openpyxl:
            raise UserError(_('openpyxl is required for Excel export.'))
        wb = openpyxl.Workbook()
        self._make_comm_summary_excel(wb, lines)

        ws = wb.create_sheet('WPS Bank File')
        thin = Border(left=Side('thin'), right=Side('thin'),
                      top=Side('thin'), bottom=Side('thin'))
        bold = Font(bold=True, size=11)
        hdr_fill = PatternFill('solid', fgColor='C6EFCE')

        cic = bank.x_wps_cic_number or ''
        debit = bank.x_wps_debit_account or ''
        mol = bank.x_wps_mol_id or ''

        ws.cell(1, 1, 'CIC').font = bold
        ws.cell(1, 2, cic)
        ws.cell(2, 1, 'Debit Account:').font = bold
        ws.cell(2, 2, debit)
        ws.cell(3, 1, 'MOL ID').font = bold
        ws.cell(3, 2, mol)
        ws.cell(4, 1, 'Payment Purpose').font = bold
        ws.cell(4, 2, 'Commissions')

        en_headers = [
            'Bank Name', 'Account Number', 'Employee Name', 'Employee Number',
            'National ID', 'Amount (SAR)', 'Basic', 'HRA',
            'Other Earnings', 'Deductions', 'Department',
        ]
        for ci, h in enumerate(en_headers, 1):
            c = ws.cell(6, ci, h)
            c.font = bold
            c.fill = hdr_fill
            c.border = thin
            c.alignment = Alignment(horizontal='center')

        ri = 7
        for line in lines.sorted(lambda s: s.employee_id.name or ''):
            amt = line.net_payable
            if not amt:
                continue
            emp = line.employee_id.sudo()
            emp_bank = getattr(emp, 'x_salary_bank_account_id', False)
            row = [
                emp_bank.bank_id.name if emp_bank and emp_bank.bank_id else '',
                emp_bank.acc_number if emp_bank else '',
                emp.name or '',
                emp.barcode or '',
                emp.identification_id or '',
                amt,
                line.earnings,
                0.0,  # HRA not applicable in commissions
                line.earnings,
                line.loan_offset,
                emp.department_id.name if emp.department_id else '',
            ]
            for ci, v in enumerate(row, 1):
                ws.cell(ri, ci, v).border = thin
            ri += 1

        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb['Sheet']
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # --- TXT generation (Kawthar format) -----------------------------------

    def _make_kawthar_txt(self, bank, lines):
        """Generate Kawthar fixed-width 194-char TXT file for commissions."""
        lines = []
        op = (self.operation_code or '2')
        vd = (self.value_date or fields.Date.context_today(self))
        vd_str = vd.strftime('%Y%m%d') if hasattr(vd, 'strftime') else str(vd).replace('-', '')

        for line in lines.sorted(lambda s: s.employee_id.name or ''):
            amt_halala = int(round((line.net_payable or 0.0) * 100))
            if amt_halala <= 0:
                continue
            emp = line.employee_id.sudo()
            emp_bank = getattr(emp, 'x_salary_bank_account_id', False)
            basic_halala = int(round((line.earnings or 0.0) * 100))

            barcode = (emp.barcode or '').ljust(12)[:12]
            cic = (bank.x_wps_cic_number or '').ljust(10)[:10]
            card_no = (emp_bank.acc_number if emp_bank else '').ljust(14)[:14]
            emp_name = (emp.name or '').ljust(50)[:50]
            nat_id = (emp.identification_id or '').ljust(10)[:10]
            net_str = str(amt_halala).zfill(15)
            basic_str = str(basic_halala).zfill(12)
            housing_str = '0' * 12
            # Everything earned this month, not just the manual lines plus
            # the driver commission. The old expression silently omitted
            # location allowance, sales, collection and combined, so the
            # breakdown never reconciled with the NET it sits next to.
            # line.total covers every contribution, present and future.
            other_str = str(int(round((line.earnings or 0.0) * 100))).zfill(12)
            ded_str = str(int(round((line.loan_offset or 0.0) * 100))).zfill(12)

            row = (
                barcode + cic + card_no + emp_name + nat_id
                + net_str + vd_str + op + '0' * 6 + ' ' * 20
                + basic_str + housing_str + other_str + ded_str
            )
            assert len(row) == 194, f"Row length {len(row)} != 194"
            lines.append(row)

        return '\n'.join(lines)

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------

    def action_export(self):
        self.ensure_one()
        mode = self.export_mode
        handlers = {
            'all_excel':     self._export_all_excel,
            'all_txt':       self._export_all_txt,
            'specific_excel': self._export_specific_excel,
            'specific_txt':  self._export_specific_txt,
        }
        return handlers[mode]()

    def _export_all_excel(self):
        groups = self._group_and_validate(require_type=None)
        files = []
        bl = self._batch_label()
        for bank, lines in groups.items():
            label = self._bank_label(bank)
            if bank.x_file_type in ('wps', 'kawthar'):
                data = self._make_wps_excel(bank, lines)
                files.append(('Commissions_%s_%s.xlsx' % (bl, label), data))
        if not files:
            raise UserError(_('No Excel files could be generated.'))
        return self._bundle_and_download(files)

    def _export_all_txt(self):
        groups = self._group_and_validate(require_type='kawthar')
        files = []
        bl = self._batch_label()
        vd = (self.value_date or fields.Date.context_today(self)
              ).strftime('%Y%m%d')
        for bank, lines in groups.items():
            label = self._bank_label(bank)
            data = self._make_kawthar_txt(bank, lines).encode('utf-8')
            files.append(('Commissions_%s_%s_%s.txt' % (bl, label, vd), data))
        if not files:
            raise UserError(_('No text files could be generated.'))
        return self._bundle_and_download(files)

    def _export_specific_excel(self):
        if not self.bank_account_id:
            raise UserError(_('Please select a bank account.'))
        bank = self.bank_account_id
        groups = self._group_and_validate()
        lines = groups.get(bank)
        if not lines:
            raise UserError(_('No lines are assigned to the selected bank.'))
        bl = self._batch_label()
        label = self._bank_label(bank)
        data = self._make_wps_excel(bank, lines)
        return self._bundle_and_download(
            [('Commissions_%s_%s.xlsx' % (bl, label), data)])

    def _export_specific_txt(self):
        if not self.bank_account_id:
            raise UserError(_('Please select a bank account.'))
        if not self.value_date:
            raise UserError(_('Please set a value date for the text file.'))
        bank = self.bank_account_id
        groups = self._group_and_validate()
        lines = groups.get(bank)
        if not lines:
            raise UserError(_('No lines are assigned to the selected bank.'))
        bl = self._batch_label()
        label = self._bank_label(bank)
        vd = self.value_date.strftime('%Y%m%d')
        data = self._make_kawthar_txt(bank, lines).encode('utf-8')
        return self._bundle_and_download(
            [('Commissions_%s_%s_%s.txt' % (bl, label, vd), data)])

