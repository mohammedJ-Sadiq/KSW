"""KSW Sales/Collection Commission Import Wizard.

Reads the accountant's monthly Excel reports and populates
``ksw.sales.commission.line.achieved_sales`` /
``ksw.sales.commission.line.achieved_collection`` /
``ksw.sales.commission.line.target_collection`` automatically.

Expected file formats (column index = 0-based)
----------------------------------------------
**Sales file** (e.g. ``Sales March 2026.xlsx``):
  Col 0: Account number
  Col 1: Customer name
  Col 2: Sales amount (VAT-inclusive)         — ignored
  Col 3: Pre-tax sales (المبيعات قبل الضريبة) → ``achieved_sales``
  Col 4: Salesman name (البائع)               → employee key

**Collection file** (e.g. ``Collection March 2026.xlsx``):
  Col 0: Account number
  Col 1: Customer name
  Col 2: Balance                              — ignored
  Col 3: Aging                                — ignored
  Col 4: Target                               → ``target_collection``
  Col 5: Amount collected (المحصل)            → ``achieved_collection``
  Col 6: Collection rep name (مندوب التحصيل)  → employee key
  Col 7: Collection %                         — ignored

Row detection
-------------
A row is considered a data row when col 0 is non-empty AND the name
column for that file is a non-empty string. Subtotal/grand-total rows
(blank account number) and ``#DIV/0!`` rows are skipped automatically.
A bad numeric cell logs a warning instead of aborting the import.

Name matching
-------------
For each name found in the file the wizard looks for an ``hr.employee``
record where ``x_commission_import_name`` equals the name (exact,
case-insensitive).  If not found, it falls back to matching on
``hr.employee.name``.  Unmatched names are listed in the chatter.
"""
import base64
import io
import logging

from markupsafe import Markup, escape

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


class KswSalesCommissionImportWizard(models.TransientModel):
    _name = 'ksw.sales.commission.import.wizard'
    _description = 'KSW Sales/Collection Commission Import Wizard'

    sheet_id = fields.Many2one(
        'ksw.sales.commission.sheet',
        required=True, ondelete='cascade',
        string='Commission Sheet',
    )
    sales_file = fields.Binary(string='Sales Excel File')
    sales_filename = fields.Char(string='Sales File Name')
    collection_file = fields.Binary(string='Collection Excel File')
    collection_filename = fields.Char(string='Collection File Name')
    collection_exclude_vat = fields.Boolean(
        string='Exclude VAT (÷ 1.15)',
        default=False,
        help='When checked, both the Target and Collected amounts from the '
             'collection file are divided by 1.15 to remove the 15% VAT '
             'before being written to the commission line.',
    )

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('openpyxl is required for Excel import.'))
        if not self.sales_file and not self.collection_file:
            raise UserError(_(
                'Please upload at least one Excel file '
                '(Sales and/or Collection).'))

        sheet = self.sheet_id
        if sheet.state != 'draft':
            raise UserError(_(
                'The commission sheet must be in draft state to import data.'))

        # -- parse files -------------------------------------------------
        # sales_data: {salesman_lower: {'total': float, 'by_customer': {cust_lower: float}}}
        # collection_totals: {rep_lower: {'collected': float, 'target': float}}
        sales_data = {}
        collection_totals = {}

        if self.sales_file:
            sales_data = self._parse_sales_file(self.sales_file)
            if not sales_data:
                raise UserError(_(
                    'No data rows could be read from the Sales file. '
                    'Check that the column layout matches the expected '
                    'format (Account · Customer · Sales · Pre-tax · '
                    'Salesman).'))
        if self.collection_file:
            collection_totals = self._parse_collection_file(
                self.collection_file,
                exclude_vat=self.collection_exclude_vat)
            if not collection_totals:
                raise UserError(_(
                    'No data rows could be read from the Collection file. '
                    'Check that the column layout matches the expected '
                    'format (Account · Customer · Balance · Aging · '
                    'Target · Collected · Rep · %).'))

        # -- build employee map ------------------------------------------
        all_names_lower = set(sales_data) | set(collection_totals)
        emp_map = self._build_employee_map(all_names_lower)
        unmatched = sorted(n for n in all_names_lower if n not in emp_map)

        # -- re-key by employee and delegate the upsert to the sheet -----
        # (shared with action_pull_from_bas() on ksw.sales.commission.sheet)
        sales_by_employee = {
            emp_map[name]: data
            for name, data in sales_data.items() if name in emp_map
        }
        collection_by_employee = {
            emp_map[name]: data
            for name, data in collection_totals.items() if name in emp_map
        }
        collection_grand_totals = None
        if self.collection_file:
            collection_grand_totals = (
                sum(v['collected'] for v in collection_totals.values()),
                sum(v['target'] for v in collection_totals.values()),
            )

        imported_lines = sheet._apply_commission_data(
            sales_by_employee, collection_by_employee,
            collection_grand_totals=collection_grand_totals)

        # -- chatter summary ----------------------------------------------
        self._post_import_summary(sheet, imported_lines, unmatched)

        return {
            'type': 'ir.actions.act_window',
            'res_id': sheet.id,
            'view_mode': 'form',
            'target': 'current',
        }

    # ------------------------------------------------------------------
    # Parsers
    # ------------------------------------------------------------------
    @staticmethod
    def _is_data_row(row, name_col):
        """Return True for a genuine data row.

        A data row has a non-empty col 0 (account number) AND a
        non-empty string in the name column. Subtotals, blanks, and
        ``#DIV/0!`` rows fail one of these conditions.
        """
        if len(row) <= name_col:
            return False
        acc = row[0]
        if acc is None:
            return False
        if isinstance(acc, str) and not acc.strip():
            return False
        name = row[name_col]
        if name is None:
            return False
        if not isinstance(name, str):
            return False
        return name.strip() != ''

    @staticmethod
    def _safe_float(value):
        """Coerce a cell value to float; return None if not coercible.

        Accepts int/float as-is, strips commas/spaces from strings,
        treats ``#DIV/0!`` and other error markers as None.
        """
        if value is None:
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.strip().replace(',', '').replace(' ', '')
            if not cleaned or cleaned.startswith('#'):
                return None
            try:
                return float(cleaned)
            except ValueError:
                return None
        return None

    def _parse_sales_file(self, file_data):
        """Return per-salesman sales data from the Excel file.

        Return format::

            {
                salesman_name_lower: {
                    'total': float,
                    'by_customer': {customer_name_lower: float},
                    'by_account':  {account_number_lower: float},
                },
                ...
            }

        ``by_account`` uses col 0 (account number) as the key —
        this is the most reliable matching key for split buckets.
        ``by_customer`` uses col 1 (customer name) as a fallback.

        See module docstring for the column contract.
        """
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(file_data)), data_only=True)
        ws = wb.active
        result = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if not self._is_data_row(row, name_col=4):
                continue
            salesman = str(row[4]).strip()
            account = str(row[0]).strip() if row[0] is not None else ''
            customer = str(row[1]).strip() if row[1] is not None else ''
            amount = self._safe_float(row[3])
            if amount is None:
                _logger.warning(
                    'KSW import: skipping non-numeric pre-tax sales '
                    'in row %d (salesman=%r, value=%r)',
                    i + 1, salesman, row[3])
                amount = 0.0
            key = salesman.lower()
            bucket = result.setdefault(
                key, {'total': 0.0, 'by_customer': {}, 'by_account': {}})
            bucket['total'] += amount
            if account:
                acc_key = account.lower()
                bucket['by_account'][acc_key] = (
                    bucket['by_account'].get(acc_key, 0.0) + amount
                )
            if customer:
                cust_key = customer.lower()
                bucket['by_customer'][cust_key] = (
                    bucket['by_customer'].get(cust_key, 0.0) + amount
                )
        return result

    def _parse_collection_file(self, file_data, exclude_vat=False):
        """Return ``{rep_name_lower: {'collected': float, 'target': float}}``
        from the collection Excel file.

        When ``exclude_vat=True`` both Target and Collected amounts are
        divided by 1.15 to strip the 15% VAT before being returned.

        See module docstring for the column contract.
        """
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(file_data)), data_only=True)
        ws = wb.active
        vat_divisor = 1.15 if exclude_vat else 1.0
        totals = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue  # header
            if not self._is_data_row(row, name_col=6):
                continue
            rep = str(row[6]).strip()
            collected = self._safe_float(row[5])
            if collected is None:
                _logger.warning(
                    'KSW import: skipping non-numeric collected amount '
                    'in row %d (rep=%r, value=%r)',
                    i + 1, rep, row[5])
                collected = 0.0
            target = self._safe_float(row[4])
            if target is None:
                target = 0.0
            key = rep.lower()
            bucket = totals.setdefault(
                key, {'collected': 0.0, 'target': 0.0})
            bucket['collected'] += collected / vat_divisor
            bucket['target'] += target / vat_divisor
        return totals

    # ------------------------------------------------------------------
    # Employee matching
    # ------------------------------------------------------------------
    def _build_employee_map(self, names_lower):
        """Return {name_lower: hr.employee} for as many names as possible.

        Priority:
          1. ``x_commission_import_name.lower()`` exact match
          2. ``hr.employee.name.lower()`` exact match
        """
        employees = self.env['hr.employee'].sudo().search([
            ('active', '=', True),
        ])
        result = {}

        # Build lookup by import alias first (highest priority).
        alias_map = {}
        for emp in employees:
            alias = (emp.x_commission_import_name or '').strip().lower()
            if alias:
                alias_map[alias] = emp

        # Build lookup by employee name.
        name_map = {}
        for emp in employees:
            name_map[emp.name.strip().lower()] = emp

        for name_lower in names_lower:
            if name_lower in alias_map:
                result[name_lower] = alias_map[name_lower]
            elif name_lower in name_map:
                result[name_lower] = name_map[name_lower]
            # else: unmatched — caller will log a warning

        return result

    # ------------------------------------------------------------------
    # Chatter summary
    # ------------------------------------------------------------------
    def _post_import_summary(self, sheet, imported_lines, unmatched):
        """Post a chatter note with a full import summary."""
        lines_html = sheet._format_commission_lines_html(imported_lines)

        warn_html = Markup('')
        if unmatched:
            names = Markup('').join(
                Markup('<li>{n}</li>').format(n=escape(n)) for n in unmatched
            )
            warn_html = Markup(
                '<br/><b>⚠ Unmatched salesman names (no employee found):</b>'
                '<ul>{names}</ul>'
                'Set the <i>Commission Import Name</i> field on the '
                'corresponding employee records to fix the mapping.'
            ).format(names=names)

        files_info = []
        if self.sales_filename:
            files_info.append(
                Markup('<b>Sales:</b> {f}').format(f=escape(self.sales_filename))
            )
        if self.collection_filename:
            vat_note = (
                Markup(' <i>(VAT excluded ÷1.15)</i>')
                if self.collection_exclude_vat else Markup('')
            )
            files_info.append(
                Markup('<b>Collection:</b> {f}{v}').format(
                    f=escape(self.collection_filename), v=vat_note)
            )
        files_str = Markup(' | ').join(files_info) if files_info else Markup('—')

        body = Markup(
            '<b>📥 Excel Import completed</b> ({files})<br/>'
            '<b>{n} line(s) updated/created:</b>'
            '<ul>{lines}</ul>'
            '{warn}'
        ).format(
            files=files_str,
            n=len(imported_lines),
            lines=lines_html,
            warn=warn_html,
        )
        sheet.message_post(body=body, subtype_xmlid='mail.mt_note')






