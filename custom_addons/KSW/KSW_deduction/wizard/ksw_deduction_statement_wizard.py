"""Statement of Account — an employee's deduction ledger.

The module already holds a complete subsidiary ledger: `ksw.deduction` is
the charge, `ksw.deduction.line` is the collection schedule. What it never
had was a way to *read* it as an account — opening balance, chronological
movements, running balance, closing balance — which is the shape SAP FI,
Oracle PeopleSoft/Fusion and Odoo's own Partner Ledger all converge on.

Nothing here stores a balance. Every figure is derived from the two
existing models on each run, so the statement can never drift out of step
with the records it describes.

Three movements make up the ledger:

* **charge**   — the deduction went active, dated `x_charge_date`, debit
                 of the full `amount`.
* **credit**   — one installment was settled, dated `x_settlement_date`
                 (when the money actually moved, *not* the month it was
                 scheduled for), credit of the line amount.
* **write-off** — the deduction was cancelled and its still-pending
                 installments skipped. Without crediting those back out,
                 a cancelled loan would carry its unrecovered balance on
                 the statement forever.

This holds together because every settlement route in the module preserves
`Σ(pending + paid) == deduction.amount` (`_validate_installments_total`),
so *charges − credits* really is what the employee still owes. The
`test_closing_matches_outstanding_total` test pins that against the
`x_deduction_outstanding_total` figure shown on the employee form.

Access: `_deductions()` searches `ksw.deduction` **as the calling user**,
so the module's existing record rules scope each role for free — the
accounting data-entry team sees no loans, loan approvers see loans, and so
on. Only the installment lines are then read with `sudo()`, which is safe
because parent access has already been established and is necessary
because `ksw.deduction.line` carries no record rules of its own (and no
ACL row at all for the GM / accounting approver groups).
"""
import base64
import io

from dateutil.relativedelta import relativedelta

from odoo import _, api, fields, models
from odoo.exceptions import AccessError, UserError, ValidationError

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover - environment guard
    openpyxl = None


# Ordering of movements that land on the same day. A deduction has to be
# charged before it can be collected, and it can only be written off once
# nothing more will be collected against it.
_KIND_ORDER = {'charge': 0, 'credit': 1, 'writeoff': 2}

# Row kinds that carry a running balance rather than a movement.
_BALANCE_KINDS = ('opening', 'subtotal', 'closing')


class KswDeductionStatementWizard(models.TransientModel):
    _name = 'ksw.deduction.statement.wizard'
    _description = 'Deduction Statement of Account'

    employee_id = fields.Many2one(
        'hr.employee', string='Employee', required=True,
        help='The employee whose deduction account is being stated.',
    )
    date_from = fields.Date(
        string='From',
        help='Leave empty to start from the very first movement. When set, '
             'everything before it is folded into the opening balance.',
    )
    date_to = fields.Date(
        string='To', required=True,
        default=lambda s: fields.Date.context_today(s),
    )
    type_ids = fields.Many2many(
        'ksw.deduction.type', string='Deduction Types',
        help='Leave empty for every type. Pick one to state a single '
             'account — a loan, the penalties, the salary advances.',
    )
    group_by_type = fields.Boolean(
        string='Group by Type', default=False,
        help='State each deduction type as its own account, with its own '
             'opening balance and subtotal, instead of one combined ledger.',
    )

    currency_id = fields.Many2one(
        'res.currency', default=lambda s: s.env.company.currency_id,
        readonly=True,
    )
    line_ids = fields.One2many(
        'ksw.deduction.statement.line', 'wizard_id', string='Statement Lines',
    )

    # Live summary, so the dialog answers the question before anything is
    # printed. Computed from the same `_movements()` the printed rows are
    # built from, so the two cannot disagree.
    opening_balance = fields.Monetary(compute='_compute_summary')
    total_charged = fields.Monetary(compute='_compute_summary')
    total_settled = fields.Monetary(compute='_compute_summary')
    total_written_off = fields.Monetary(compute='_compute_summary')
    closing_balance = fields.Monetary(compute='_compute_summary')
    overdue_amount = fields.Monetary(
        compute='_compute_summary',
        help='Of the closing balance, how much was scheduled for a month '
             'that has already passed and is still pending.',
    )
    movement_count = fields.Integer(compute='_compute_summary')

    # A statement that quietly omits rows is worse than no statement.
    # Several roles are scoped to a subset of the types — the GM to his
    # own departments' loans, the accounting data-entry team to non-loans
    # — so say so on the face of the document rather than presenting a
    # truncated ledger as complete.
    out_of_scope_count = fields.Integer(
        compute='_compute_summary',
        help='Deductions this employee has that your access rights do not '
             'let you see. They are excluded from every figure here.',
    )
    # `ksw.deduction.amount` has no server-side constraint tying it to its
    # installment lines (`_validate_installments_total` only runs when a
    # write touches `line_ids`), so a raw write can desync the two. The
    # statement reports that rather than papering over it.
    reconciliation_warning = fields.Char(compute='_compute_summary')

    # ------------------------------------------------------------------
    # Defaults & validation
    # ------------------------------------------------------------------

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        # Opened from a deduction form or an employee form — carry the
        # employee across so the user does not retype it.
        ctx = self.env.context
        if not res.get('employee_id'):
            if ctx.get('active_model') == 'ksw.deduction' and ctx.get('active_id'):
                deduction = self.env['ksw.deduction'].browse(ctx['active_id'])
                res['employee_id'] = deduction.employee_id.id
            elif ctx.get('active_model') == 'hr.employee' and ctx.get('active_id'):
                res['employee_id'] = ctx['active_id']
        if 'date_from' in fields_list and not res.get('date_from'):
            today = fields.Date.context_today(self)
            res['date_from'] = today.replace(month=1, day=1)
        return res

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for wiz in self:
            if wiz.date_from and wiz.date_to and wiz.date_from > wiz.date_to:
                raise ValidationError(
                    _('"From" must be on or before "To".'))

    # ------------------------------------------------------------------
    # Gathering the movements
    # ------------------------------------------------------------------

    def _deductions(self):
        """The in-scope deductions, searched AS THE USER.

        Running the search unprivileged is the whole access story: the
        thirteen `ksw.deduction` record rules already express who may see
        which types, so the statement inherits the correct scope for every
        role without restating it. A `draft` deduction is excluded because
        nothing has been charged to the employee yet.
        """
        self.ensure_one()
        domain = [
            ('employee_id', '=', self.employee_id.id),
            ('state', 'in', ('active', 'completed', 'cancelled')),
        ]
        if self.type_ids:
            domain.append(('type_id', 'in', self.type_ids.ids))
        return self.env['ksw.deduction'].search(domain)

    def _movements(self):
        """Every ledger movement in scope, oldest first.

        Returns a list of plain dicts rather than records: a movement is
        not a stored thing, and keeping it that way makes it obvious that
        the statement invents no data. Both `_compute_summary` and
        `_build_lines` consume this, which is what guarantees the screen,
        the PDF and the spreadsheet agree.
        """
        self.ensure_one()
        movements = []
        # Parent access is established by the search above; the lines are
        # then read with sudo() because ksw.deduction.line has no record
        # rules and no ACL row for several of the approver groups.
        for deduction in self._deductions().sudo():
            ref = deduction.name or ''
            # --- the charge -------------------------------------------
            # `x_charge_date` is stamped at activation. Older records
            # predating the field fall back to the month the schedule
            # starts, which is the closest honest approximation.
            charge_date = deduction.x_charge_date or deduction.start_month
            if charge_date:
                movements.append({
                    'date': charge_date,
                    'kind': 'charge',
                    'deduction': deduction,
                    'label': _('%(type)s granted — %(n)s installment(s)',
                               type=deduction.type_id.name or '',
                               n=deduction.installments),
                    'ref': ref,
                    'debit': deduction.amount,
                    'credit': 0.0,
                    'sort': (charge_date, _KIND_ORDER['charge'], ref, 0),
                })
            # --- the collections --------------------------------------
            for line in deduction.line_ids.filtered(
                    lambda l: l.state == 'paid'):
                settled_on = self._settlement_date(line)
                if not settled_on:
                    continue
                movements.append({
                    'date': settled_on,
                    'kind': 'credit',
                    'deduction': deduction,
                    'label': self._credit_label(line),
                    'ref': ref,
                    'debit': 0.0,
                    'credit': line.amount,
                    'sort': (settled_on, _KIND_ORDER['credit'], ref,
                             line.sequence),
                })
            # --- the write-off ----------------------------------------
            # A cancelled deduction keeps its charge but will never
            # collect the rest. Credit the shortfall back out or the
            # account never closes.
            #
            # The figure is `amount − Σpaid`, NOT the sum of the skipped
            # lines. Those are not the same: adding a manual paid line
            # forces the accountant to skip a compensating pending line
            # first (`_validate_installments_total` refuses otherwise —
            # see the pattern in tests/test_managed_by.py), so an *active*
            # deduction can already carry skipped lines. Cancelling then
            # skips the remaining pending ones too, and crediting every
            # skipped line would double-count that earlier adjustment and
            # drive the statement negative.
            if deduction.state == 'cancelled':
                paid_total = sum(deduction.line_ids.filtered(
                    lambda l: l.state == 'paid').mapped('amount'))
                shortfall = deduction.amount - paid_total
                if deduction.currency_id.compare_amounts(shortfall, 0) > 0:
                    off_date = (deduction.x_writeoff_date
                                or deduction.write_date.date())
                    movements.append({
                        'date': off_date,
                        'kind': 'writeoff',
                        'deduction': deduction,
                        'label': _('Cancelled — uncollected balance '
                                   'written off'),
                        'ref': ref,
                        'debit': 0.0,
                        'credit': shortfall,
                        'sort': (off_date, _KIND_ORDER['writeoff'], ref, 0),
                    })
        movements.sort(key=lambda m: m['sort'])
        return movements

    @staticmethod
    def _settlement_date(line):
        """When a settled installment's money actually moved.

        `x_settlement_date` is stamped by every settlement route in the
        codebase, but `state` is not in `_INSTALLMENT_EDIT_KEYS`, so
        `ksw.deduction.line.write()` early-returns on a bare
        `write({'state': 'paid'})` — an officer can flip a line over RPC
        with nothing to intercept it, and historical rows predate the
        field entirely. Rather than drop such a row from the ledger (which
        would silently understate the debt), fall back through the
        evidence that does exist. Called on sudo lines, so reading
        `payslip_id` needs no payroll access.
        """
        return (
            line.x_settlement_date
            or line.manual_date
            or line.payslip_id.date_to
            or line.period_date
        )

    def _credit_label(self, line):
        """How a settled installment describes itself on the statement."""
        position = _('Installment %(label)s', label=line.display_name)
        settlement = line.settlement_label
        if settlement:
            return '%s — %s' % (position, settlement)
        return position

    def _split_movements(self, movements):
        """Split into (before the period, inside it), honouring `date_to`.

        Anything after `date_to` is not on this statement at all — a
        statement cannot report movements that had not happened yet as of
        the date it is dated.
        """
        self.ensure_one()
        before, inside = [], []
        for mv in movements:
            if mv['date'] > self.date_to:
                continue
            if self.date_from and mv['date'] < self.date_from:
                before.append(mv)
            else:
                inside.append(mv)
        return before, inside

    @staticmethod
    def _net(movements):
        return sum(m['debit'] - m['credit'] for m in movements)

    # ------------------------------------------------------------------
    # Summary
    # ------------------------------------------------------------------

    @api.depends_context('uid')
    @api.depends('employee_id', 'date_from', 'date_to', 'type_ids')
    def _compute_summary(self):
        for wiz in self:
            if not wiz.employee_id or not wiz.date_to:
                wiz.opening_balance = wiz.total_charged = 0.0
                wiz.total_settled = wiz.total_written_off = 0.0
                wiz.closing_balance = wiz.overdue_amount = 0.0
                wiz.movement_count = wiz.out_of_scope_count = 0
                wiz.reconciliation_warning = False
                continue
            before, inside = wiz._split_movements(wiz._movements())
            opening = wiz._net(before)
            wiz.opening_balance = opening
            wiz.total_charged = sum(m['debit'] for m in inside)
            wiz.total_settled = sum(
                m['credit'] for m in inside if m['kind'] == 'credit')
            wiz.total_written_off = sum(
                m['credit'] for m in inside if m['kind'] == 'writeoff')
            wiz.closing_balance = opening + wiz._net(inside)
            wiz.movement_count = len(inside)
            wiz.overdue_amount = wiz._overdue_amount()
            wiz.out_of_scope_count = wiz._out_of_scope_count()
            wiz.reconciliation_warning = wiz._reconciliation_warning()

    def _out_of_scope_count(self):
        """How many of this employee's deductions the user cannot see.

        The same domain, run once as the user and once with `sudo()`. The
        difference is what the record rules withheld — disclosed on the
        statement so a partially-scoped role knows the ledger in front of
        them is not the whole account.
        """
        self.ensure_one()
        domain = [
            ('employee_id', '=', self.employee_id.id),
            ('state', 'in', ('active', 'completed', 'cancelled')),
        ]
        if self.type_ids:
            domain.append(('type_id', 'in', self.type_ids.ids))
        Deduction = self.env['ksw.deduction']
        return (Deduction.sudo().search_count(domain)
                - Deduction.search_count(domain))

    def _reconciliation_warning(self):
        """Flag deductions whose header amount no longer matches their lines.

        `_validate_installments_total` only fires when a write touches
        `line_ids`, so `write({'amount': ...})` on an active deduction
        desyncs the two silently. When that has happened the charge row
        (the header amount) and the outstanding figure (the sum of the
        lines) genuinely disagree, and the statement should say which
        documents to look at rather than present a balance it cannot
        stand behind.
        """
        self.ensure_one()
        broken = []
        for deduction in self._deductions().sudo():
            if deduction.state not in ('active', 'completed'):
                continue
            counted = sum(deduction.line_ids.filtered(
                lambda l: l.state in ('pending', 'paid')).mapped('amount'))
            if deduction.currency_id.compare_amounts(
                    counted, deduction.amount) != 0:
                broken.append(deduction.name or str(deduction.id))
        if not broken:
            return False
        return _(
            'The total amount does not match the installment schedule on: '
            '%(refs)s. The balance below is derived from the recorded '
            'amounts and should be reconciled before the statement is '
            'relied on.',
            refs=', '.join(broken),
        )

    def _overdue_amount(self):
        """Of what is still owed, how much is already late.

        A memo figure beside the closing balance, in the spirit of the
        aging panel SAP puts next to an account statement. It reuses the
        module's own definition of overdue (`x_is_overdue`): pending, in a
        month that has already started.
        """
        self.ensure_one()
        current_month = fields.Date.today().replace(day=1)
        total = 0.0
        for deduction in self._deductions().sudo():
            if deduction.state != 'active':
                continue
            for line in deduction.line_ids:
                if (line.state == 'pending' and line.period_date
                        and line.period_date < current_month):
                    total += line.amount
        return total

    # ------------------------------------------------------------------
    # Materialising the rows
    # ------------------------------------------------------------------

    def _build_lines(self):
        """Turn the movements into the rows every output renders.

        Materialised rather than recomputed per output: the list view, the
        QWeb template and the spreadsheet all read these same records, so
        it is structurally impossible for the three to show different
        numbers.
        """
        self.ensure_one()
        self.line_ids.unlink()
        before, inside = self._split_movements(self._movements())
        vals = (self._grouped_rows(before, inside) if self.group_by_type
                else self._flat_rows(before, inside))
        for index, row in enumerate(vals):
            row['wizard_id'] = self.id
            row['sequence'] = index + 1
            row.setdefault('currency_id', self.currency_id.id)
        return self.env['ksw.deduction.statement.line'].create(vals)

    def _flat_rows(self, before, inside):
        """One combined account: opening, movements, closing."""
        balance = self._net(before)
        rows = [self._balance_row(
            'opening', _('Opening balance'), self.date_from, balance)]
        for mv in inside:
            balance += mv['debit'] - mv['credit']
            rows.append(self._movement_row(mv, balance))
        rows.append(self._balance_row(
            'closing', _('Closing balance'), self.date_to, balance))
        return rows

    def _grouped_rows(self, before, inside):
        """One account per deduction type, then a grand total.

        Each type carries its own opening balance and subtotal, which is
        what makes "the parts sum to the whole" checkable by eye — and is
        how the same report reads when filtered to a single type.
        """
        types = self.env['ksw.deduction.type']
        for mv in before + inside:
            types |= mv['deduction'].type_id
        rows = []
        grand_opening = grand_closing = 0.0
        for ded_type in types.sorted(lambda t: (t.sequence, t.id)):
            t_before = [m for m in before
                        if m['deduction'].type_id == ded_type]
            t_inside = [m for m in inside
                        if m['deduction'].type_id == ded_type]
            balance = self._net(t_before)
            grand_opening += balance
            rows.append(self._balance_row(
                'opening',
                _('Opening balance — %(type)s', type=ded_type.name or ''),
                self.date_from, balance, ded_type=ded_type))
            for mv in t_inside:
                balance += mv['debit'] - mv['credit']
                rows.append(self._movement_row(mv, balance))
            grand_closing += balance
            rows.append(self._balance_row(
                'subtotal',
                _('Subtotal — %(type)s', type=ded_type.name or ''),
                self.date_to, balance, ded_type=ded_type))
        rows.insert(0, self._balance_row(
            'opening', _('Opening balance — all types'),
            self.date_from, grand_opening))
        rows.append(self._balance_row(
            'closing', _('Closing balance — all types'),
            self.date_to, grand_closing))
        return rows

    def _balance_row(self, kind, label, date, balance, ded_type=None):
        return {
            'row_kind': kind,
            'date': date,
            'label': label,
            'type_id': ded_type.id if ded_type else False,
            'debit': 0.0,
            'credit': 0.0,
            'balance': balance,
        }

    def _movement_row(self, mv, balance):
        return {
            'row_kind': mv['kind'],
            'date': mv['date'],
            'ref': mv['ref'],
            'deduction_id': mv['deduction'].id,
            'type_id': mv['deduction'].type_id.id,
            'label': mv['label'],
            'debit': mv['debit'],
            'credit': mv['credit'],
            'balance': balance,
        }

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------

    def action_view(self):
        """Open the statement on screen.

        The context deliberately carries no `default_*`: the **New
        Statement** button on the resulting list opens a blank dialog, and
        a leftover default here would pre-fill it.
        """
        self.ensure_one()
        self._build_lines()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Statement of Account — %(name)s',
                      name=self.employee_id.name or ''),
            'res_model': 'ksw.deduction.statement.line',
            'view_mode': 'list',
            'domain': [('wizard_id', '=', self.id)],
            'context': {'create': False, 'edit': False},
            'target': 'current',
        }

    def action_print(self):
        self.ensure_one()
        self._build_lines()
        return self.env.ref(
            'KSW_deduction.action_report_deduction_statement'
        ).report_action(self)

    def action_export_xlsx(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_(
                'The openpyxl library is required to export to Excel. '
                'Print the PDF instead, or ask an administrator to '
                'install it.'))
        self._build_lines()
        content = self._build_workbook()
        filename = '%s - Statement of Account - %s.xlsx' % (
            self.employee_id.name or '',
            fields.Date.to_string(self.date_to),
        )
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(content),
            'mimetype': 'application/vnd.openxmlformats-officedocument.'
                        'spreadsheetml.sheet',
            # Deliberately unattached. `ir.attachment` resolves access
            # through the referenced record, and this wizard is transient
            # — once it is vacuumed, an attachment pointing at it becomes
            # unreachable even for the person who just downloaded it. With
            # no res_model the creator keeps access.
            'res_model': False,
            'res_id': False,
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'new',
        }

    # ------------------------------------------------------------------
    # Rendering
    # ------------------------------------------------------------------

    def get_report_data(self):
        """Everything the QWeb template needs, in one call.

        Follows the shape `KSW_attendance_report` established: the
        template unpacks this dict and renders, holding no logic of its
        own.
        """
        self.ensure_one()
        if not self.line_ids:
            self._build_lines()
        return {
            'employee': self.employee_id,
            'company': self.env.company,
            'date_from': self.date_from,
            'date_to': self.date_to,
            'types': self.type_ids,
            'lines': self.line_ids.sorted('sequence'),
            'currency': self.currency_id,
            'opening_balance': self.opening_balance,
            'total_charged': self.total_charged,
            'total_settled': self.total_settled,
            'total_written_off': self.total_written_off,
            'closing_balance': self.closing_balance,
            'overdue_amount': self.overdue_amount,
            'printed_by': self.env.user,
            'printed_on': fields.Date.context_today(self),
        }

    def _build_workbook(self):
        """The same rows as the screen and the PDF, as a spreadsheet."""
        self.ensure_one()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Statement'
        bold = Font(bold=True)
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4F6228')
        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        money = '#,##0.00'

        ws['A1'] = _('Statement of Account')
        ws['A1'].font = Font(bold=True, size=14)
        meta = [
            (_('Employee'), self.employee_id.name or ''),
            (_('Period'), '%s → %s' % (
                fields.Date.to_string(self.date_from) if self.date_from
                else _('inception'),
                fields.Date.to_string(self.date_to))),
            (_('Deduction Types'), ', '.join(
                self.type_ids.mapped('name')) or _('All types')),
            (_('Currency'), self.currency_id.name or ''),
        ]
        row = 3
        for label, value in meta:
            ws.cell(row=row, column=1, value=label).font = bold
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        headers = [_('Date'), _('Reference'), _('Type'), _('Description'),
                   _('Charge'), _('Collected'), _('Balance')]
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        row += 1

        for line in self.line_ids.sorted('sequence'):
            is_balance = line.row_kind in _BALANCE_KINDS
            values = [
                line.date or '',
                line.ref or '',
                line.type_id.name or '',
                line.label or '',
                line.debit or None,
                line.credit or None,
                line.balance,
            ]
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col >= 5:
                    cell.number_format = money
                    cell.alignment = Alignment(horizontal='right')
                if is_balance:
                    cell.font = bold
            row += 1

        widths = [12, 14, 20, 52, 14, 14, 16]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)].width = width
        ws.freeze_panes = ws.cell(row=row - len(self.line_ids), column=1)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class KswDeductionStatementLine(models.TransientModel):
    _name = 'ksw.deduction.statement.line'
    _description = 'Deduction Statement Line'
    _order = 'sequence, id'

    wizard_id = fields.Many2one(
        'ksw.deduction.statement.wizard', required=True, ondelete='cascade',
        index=True,
    )
    sequence = fields.Integer(default=1)
    row_kind = fields.Selection([
        ('opening', 'Opening Balance'),
        ('charge', 'Charge'),
        ('credit', 'Collected'),
        ('writeoff', 'Written Off'),
        ('subtotal', 'Subtotal'),
        ('closing', 'Closing Balance'),
    ], required=True)
    date = fields.Date()
    ref = fields.Char(string='Reference')
    deduction_id = fields.Many2one('ksw.deduction', string='Deduction')
    type_id = fields.Many2one('ksw.deduction.type', string='Type')
    label = fields.Char(string='Description')
    debit = fields.Monetary(string='Charge')
    credit = fields.Monetary(string='Collected')
    balance = fields.Monetary(string='Balance')
    currency_id = fields.Many2one('res.currency', readonly=True)

    # A balance row is not a transaction; the list view uses this to bold
    # it and to hide the drill-down button that would have nowhere to go.
    is_balance_row = fields.Boolean(compute='_compute_is_balance_row')

    @api.depends('row_kind')
    def _compute_is_balance_row(self):
        for line in self:
            line.is_balance_row = line.row_kind in _BALANCE_KINDS

    def _compute_display_name(self):
        for line in self:
            line.display_name = line.label or dict(
                self._fields['row_kind'].selection).get(line.row_kind, '')

    def action_reopen_wizard(self):
        """Open a new, blank statement dialog.

        Exactly what Deductions → Reports → Statement of Account gives you,
        just without leaving the statement you are looking at. It returns
        **the menu's own action**, read from its xml id, rather than a
        hand-built dict — so the two entry points cannot drift apart, and
        anything later added to the action (a domain, a default, a renamed
        view) is picked up here for free.

        The context is blanked deliberately. Odoo evaluates a returned
        action's context against the calling one, so any `default_*` in
        scope would silently pre-fill the dialog — which is the opposite of
        a blank form.

        **Deliberately NOT `@api.model`**, though it uses no records. A
        button always sends ids, and `call_kw` only strips them for
        non-model methods:

            if getattr(method, '_api_model', False):
                recs = model              # args NOT stripped
            else:
                ids, args = args[0], args[1:]

        so `@api.model` leaves the id list in `args` and the call arrives as
        `method(recs, [ids])` — `TypeError: takes 1 positional argument but
        2 were given`, every time, for any button. A plain method that never
        reads `self` is what the header needs: with `display="always"` it is
        clicked with nothing selected, so `self` is an empty recordset.
        """
        action = self.env['ir.actions.act_window']._for_xml_id(
            'KSW_deduction.action_ksw_deduction_statement')
        action['context'] = {}
        return action

    def action_open_deduction(self):
        """Drill from a statement row to the document behind it.

        Mirrors `ksw.deduction.line.action_open_deduction`: a user can
        legitimately see a movement whose parent their record rules will
        not open, and a readable message beats an AccessError traceback.
        """
        self.ensure_one()
        deduction = self.deduction_id
        if not deduction:
            raise UserError(_(
                'This is a balance line, not a transaction — there is no '
                'document behind it.'))
        try:
            deduction.check_access('read')
        except AccessError:
            raise UserError(_(
                'You do not have access to deduction %(name)s.',
                name=deduction.sudo().display_name))
        return {
            'type': 'ir.actions.act_window',
            'name': deduction.display_name,
            'res_model': 'ksw.deduction',
            'res_id': deduction.id,
            'view_mode': 'form',
            'target': 'current',
        }
